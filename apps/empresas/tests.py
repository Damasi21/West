import json
from datetime import timedelta
from io import BytesIO
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import (
    CadastroOmie,
    CategoriaOmie,
    ContaCorrenteOmie,
    ContaPagarOmie,
    ContaReceberOmie,
    ContaDRE,
    DepartamentoOmie,
    Empresa,
    EmpresaUsuario,
    IntegracaoOmie,
    LancamentoContaCorrenteOmie,
    ProjetoOmie,
    SincronizacaoOmie,
    TipoContaCorrenteOmie,
)
from .omie import (
    consultar_contas_correntes,
    consultar_contas_pagar,
    consultar_contas_receber,
    consultar_lancamentos_conta_corrente,
    consultar_tipos_conta_corrente,
    executar_sincronizacao_omie,
)


def arquivo_xlsx(nome_aba, cabecalho, linhas):
    workbook = Workbook()
    ws = workbook.active
    ws.title = nome_aba
    ws.append(cabecalho)
    for linha in linhas:
        ws.append(linha)
    conteudo = BytesIO()
    workbook.save(conteudo)
    return SimpleUploadedFile(
        "planilha.xlsx",
        conteudo.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


class ListaEmpresasTests(TestCase):
    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="cliente", password="senha-segura"
        )
        self.empresa_permitida = Empresa.objects.create(
            nome="Empresa Permitida Ltda",
            nome_fantasia="Empresa Permitida",
            cnpj="00.000.000/0001-01",
        )
        self.empresa_bloqueada = Empresa.objects.create(
            nome="Empresa Bloqueada Ltda",
            nome_fantasia="Empresa Bloqueada",
            cnpj="00.000.000/0001-02",
        )
        EmpresaUsuario.objects.create(
            empresa=self.empresa_permitida,
            usuario=self.usuario,
            papel=EmpresaUsuario.Papel.VISUALIZADOR,
        )

    def test_usuario_anonimo_e_redirecionado_para_login(self):
        response = self.client.get(reverse("empresas:lista"))
        self.assertRedirects(
            response,
            f"{reverse('accounts:login')}?next={reverse('empresas:lista')}",
        )

    def test_cliente_visualiza_apenas_empresa_vinculada(self):
        self.client.force_login(self.usuario)
        response = self.client.get(reverse("empresas:lista"))

        self.assertContains(response, self.empresa_permitida.nome_fantasia)
        self.assertNotContains(response, self.empresa_bloqueada.nome_fantasia)
        self.assertNotContains(response, "Configurações")
        self.assertContains(response, "MD21 BI")
        self.assertContains(response, "/media/md21_bi.png")


class ConfiguracoesEmpresasTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.administrador = User.objects.create_user(
            username="admin_oeste",
            password="senha-segura",
            is_staff=True,
        )
        self.cliente = User.objects.create_user(
            username="cliente_comum",
            password="senha-segura",
        )

    def test_administrador_visualiza_configuracoes_na_tela_inicial(self):
        self.client.force_login(self.administrador)
        response = self.client.get(reverse("empresas:lista"))

        self.assertContains(response, "Configurações")
        self.assertContains(response, reverse("empresas:cadastrar"))
        self.assertNotContains(response, "Nova empresa")
        self.assertNotContains(response, "Parâmetros")

    def test_cliente_nao_acessa_configuracoes_diretamente(self):
        self.client.force_login(self.cliente)

        response = self.client.get(reverse("empresas:configuracoes"))

        self.assertEqual(response.status_code, 403)

    def test_administrador_cadastra_empresa(self):
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse("empresas:cadastrar"),
            {
                "nome_fantasia": "Cliente Oeste",
                "nome": "Cliente Oeste Comércio Ltda",
                "cnpj": "12.345.678/0001-90",
                "ativa": "on",
            },
        )

        self.assertRedirects(response, reverse("empresas:configuracoes"))
        empresa = Empresa.objects.get(cnpj="12.345.678/0001-90")
        self.assertEqual(empresa.nome_fantasia, "Cliente Oeste")
        self.assertTrue(empresa.ativa)

    def test_cadastro_rejeita_cnpj_incompleto(self):
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse("empresas:cadastrar"),
            {
                "nome_fantasia": "Empresa Inválida",
                "nome": "Empresa Inválida Ltda",
                "cnpj": "123",
                "ativa": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Informe um CNPJ com 14 números.")
        self.assertFalse(Empresa.objects.filter(nome_fantasia="Empresa Inválida").exists())


class ParametrosOmieTests(TestCase):
    def setUp(self):
        self.administrador = get_user_model().objects.create_user(
            username="admin_parametros",
            password="senha-segura",
            is_staff=True,
        )
        self.cliente = get_user_model().objects.create_user(
            username="cliente_parametros",
            password="senha-segura",
        )
        self.empresa = Empresa.objects.create(
            nome="Empresa OMIE Ltda",
            nome_fantasia="Empresa OMIE",
            cnpj="00.000.000/0001-03",
        )

    def test_cliente_nao_acessa_parametros(self):
        self.client.force_login(self.cliente)
        response = self.client.get(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )
        self.assertEqual(response.status_code, 403)

    def test_administrador_seleciona_empresa(self):
        self.client.force_login(self.administrador)
        response = self.client.get(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
        )
        self.assertContains(response, "Credenciais exclusivas de")
        self.assertContains(response, self.empresa.nome_fantasia)

    def test_salva_credenciais_isoladas_e_criptografadas(self):
        self.client.force_login(self.administrador)
        response = self.client.post(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {
                "app_key": "minha-app-key",
                "app_secret": "segredo-super-secreto",
                "ativa": "on",
            },
        )

        self.assertEqual(response.status_code, 302)
        integracao = IntegracaoOmie.objects.get(empresa=self.empresa)
        self.assertEqual(integracao.app_key, "minha-app-key")
        self.assertNotEqual(
            integracao.app_secret_criptografado,
            "segredo-super-secreto",
        )
        self.assertEqual(integracao.obter_app_secret(), "segredo-super-secreto")

    def test_edicao_sem_novo_secret_preserva_o_existente(self):
        integracao = IntegracaoOmie(empresa=self.empresa, app_key="key-antiga")
        integracao.definir_app_secret("secret-original")
        integracao.save()
        self.client.force_login(self.administrador)

        self.client.post(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {
                "app_key": "key-nova",
                "app_secret": "",
                "ativa": "on",
            },
        )

        integracao.refresh_from_db()
        self.assertEqual(integracao.app_key, "key-nova")
        self.assertEqual(integracao.obter_app_secret(), "secret-original")


class EstruturaDRETests(TestCase):
    def setUp(self):
        self.administrador = get_user_model().objects.create_user(
            username="admin_dre",
            password="senha-segura",
            is_staff=True,
        )
        self.cliente = get_user_model().objects.create_user(
            username="cliente_dre",
            password="senha-segura",
        )
        self.empresa = Empresa.objects.create(
            nome="Empresa DRE Ltda",
            nome_fantasia="Empresa DRE",
            cnpj="00.000.000/0001-05",
        )
        self.url = reverse(
            "dashboards:dre_categorias",
            kwargs={"empresa_slug": self.empresa.slug},
        )

    def test_cards_dre_e_categorias_aparecem_separados_em_parametros(self):
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertContains(response, "<h2>DRE</h2>", html=True)
        self.assertContains(response, "<h2>Categorias</h2>", html=True)
        self.assertNotContains(response, "DRE e Categorias")
        self.assertContains(response, self.url)
        self.assertContains(
            response,
            reverse(
                "dashboards:categorias",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
        )

    def test_usuario_comum_nao_acessa_estrutura_dre(self):
        self.client.force_login(self.cliente)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_cria_conta_pai_e_conta_filha(self):
        self.client.force_login(self.administrador)
        self.client.post(
            self.url,
            {
                "nome": "Receitas operacionais",
                "tipo": "pai",
                "conta_pai": "",
                "sinal": "+",
            },
        )
        pai = ContaDRE.objects.get(nome="Receitas operacionais")

        response = self.client.post(
            self.url,
            {
                "nome": "Receita de vendas",
                "tipo": "filho",
                "conta_pai": pai.pk,
                "sinal": "+",
            },
        )

        self.assertRedirects(response, self.url)
        filha = ContaDRE.objects.get(nome="Receita de vendas")
        self.assertEqual(filha.conta_pai, pai)
        self.assertEqual(filha.nivel, 2)

    def test_conta_filha_exige_conta_pai(self):
        self.client.force_login(self.administrador)

        response = self.client.post(
            self.url,
            {
                "nome": "Conta sem grupo",
                "tipo": "filho",
                "conta_pai": "",
                "sinal": "-",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Selecione a conta pai.")
        self.assertFalse(ContaDRE.objects.filter(nome="Conta sem grupo").exists())

    def test_reordena_grupos_da_arvore(self):
        primeira = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Primeira",
            sinal="+",
            ordem=1,
        )
        segunda = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Segunda",
            sinal="=",
            ordem=2,
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse(
                "dashboards:reordenar_contas_dre",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            data=json.dumps(
                {
                    "parent_id": None,
                    "ids": [segunda.pk, primeira.pk],
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        primeira.refresh_from_db()
        segunda.refresh_from_db()
        self.assertEqual(segunda.ordem, 1)
        self.assertEqual(primeira.ordem, 2)

    def test_nao_exclui_grupo_com_conta_filha(self):
        pai = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Grupo",
            ordem=1,
        )
        ContaDRE.objects.create(
            empresa=self.empresa,
            conta_pai=pai,
            nome="Filha",
            ordem=1,
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse(
                "dashboards:excluir_conta_dre",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "conta_id": pai.pk,
                },
            )
        )

        self.assertRedirects(response, self.url)
        self.assertTrue(ContaDRE.objects.filter(pk=pai.pk).exists())

    def test_exporta_planilha_dre_com_arvore_e_listas_de_selecao(self):
        pai = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receitas",
            sinal="+",
            ordem=1,
        )
        ContaDRE.objects.create(
            empresa=self.empresa,
            conta_pai=pai,
            nome="Vendas",
            sinal="-",
            ordem=1,
        )
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:exportar_planilha_dre",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook["DRE"]
        self.assertEqual(
            [ws.cell(1, coluna).value for coluna in range(1, 4)],
            ["Nome da conta DRE", "Tipo", "Operação"],
        )
        self.assertEqual([ws["A2"].value, ws["B2"].value, ws["C2"].value], ["Receitas", "Pai", "+"])
        self.assertEqual([ws["A3"].value, ws["B3"].value, ws["C3"].value], ["Vendas", "Filho", "-"])
        self.assertTrue(ws.data_validations.dataValidation)

    def test_importa_planilha_dre_e_monta_relacao_pai_filho(self):
        self.client.force_login(self.administrador)
        planilha = arquivo_xlsx(
            "DRE",
            ("Nome da conta DRE", "Tipo", "Operação"),
            (
                ("Receitas", "Pai", "+"),
                ("Venda de serviços", "Filho", "+"),
                ("Resultado", "Pai", "="),
            ),
        )

        response = self.client.post(
            reverse(
                "dashboards:importar_planilha_dre",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {"planilha": planilha},
        )

        self.assertRedirects(response, self.url)
        receitas = ContaDRE.objects.get(empresa=self.empresa, nome="Receitas")
        venda = ContaDRE.objects.get(empresa=self.empresa, nome="Venda de serviços")
        self.assertEqual(venda.conta_pai, receitas)
        self.assertEqual(venda.ordem, 1)
        self.assertEqual(
            ContaDRE.objects.get(empresa=self.empresa, nome="Resultado").sinal,
            "=",
        )

    def test_importacao_dre_exige_confirmacao_para_sobrepor(self):
        atual = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Estrutura atual",
            ordem=1,
        )
        self.client.force_login(self.administrador)
        url = reverse(
            "dashboards:importar_planilha_dre",
            kwargs={"empresa_slug": self.empresa.slug},
        )

        response = self.client.post(
            url,
            {
                "planilha": arquivo_xlsx(
                    "DRE",
                    ("Nome da conta DRE", "Tipo", "Operação"),
                    (("Nova estrutura", "Pai", "+"),),
                )
            },
        )

        self.assertRedirects(response, self.url)
        self.assertTrue(ContaDRE.objects.filter(pk=atual.pk).exists())
        self.assertFalse(ContaDRE.objects.filter(nome="Nova estrutura").exists())

        response = self.client.post(
            url,
            {
                "sobrepor": "sim",
                "planilha": arquivo_xlsx(
                    "DRE",
                    ("Nome da conta DRE", "Tipo", "Operação"),
                    (("Nova estrutura", "Pai", "+"),),
                ),
            },
        )

        self.assertRedirects(response, self.url)
        self.assertFalse(ContaDRE.objects.filter(pk=atual.pk).exists())
        self.assertTrue(ContaDRE.objects.filter(nome="Nova estrutura").exists())


class CategoriasOmieTests(TestCase):
    def setUp(self):
        self.administrador = get_user_model().objects.create_user(
            username="admin_categorias",
            password="senha-segura",
            is_staff=True,
        )
        self.cliente = get_user_model().objects.create_user(
            username="cliente_categorias",
            password="senha-segura",
        )
        self.empresa = Empresa.objects.create(
            nome="Empresa Categorias Ltda",
            nome_fantasia="Empresa Categorias",
            cnpj="00.000.000/0001-06",
        )
        self.url = reverse(
            "dashboards:categorias",
            kwargs={"empresa_slug": self.empresa.slug},
        )
        self.conta_dre = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receita operacional",
            ordem=1,
        )
        self.categoria_pai = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01",
            descricao="Receitas",
            conta_receita=True,
        )
        self.categoria_filha = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.01",
            categoria_superior="1.01",
            descricao="Venda de mercadorias",
            conta_receita=False,
        )
        self.categoria_inativa = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.02",
            categoria_superior="1.01",
            descricao="Categoria inativa",
            conta_inativa=True,
        )

    def test_exibe_apenas_ativas_com_regras_visuais_e_seletor_na_filha(self):
        self.client.force_login(self.administrador)

        response = self.client.get(self.url)

        self.assertContains(response, self.categoria_pai.codigo)
        self.assertContains(response, self.categoria_filha.codigo)
        self.assertNotContains(response, self.categoria_inativa.descricao)
        self.assertContains(response, "is-revenue")
        self.assertContains(response, "is-expense")
        self.assertContains(response, "is-parent")
        self.assertContains(response, "Salvar associações", count=2)
        self.assertNotContains(
            response,
            f'name="conta_dre_{self.categoria_pai.pk}"',
        )
        self.assertContains(
            response,
            f'name="conta_dre_{self.categoria_filha.pk}"',
        )

    def test_salva_vinculo_da_categoria_filha_com_conta_dre(self):
        self.client.force_login(self.administrador)

        response = self.client.post(
            self.url,
            {
                f"conta_dre_{self.categoria_filha.pk}": self.conta_dre.pk,
            },
        )

        self.assertRedirects(response, self.url)
        self.categoria_filha.refresh_from_db()
        self.assertEqual(self.categoria_filha.conta_dre, self.conta_dre)

    def test_rejeita_conta_dre_de_outra_empresa(self):
        outra_empresa = Empresa.objects.create(
            nome="Outra Empresa Categorias Ltda",
            nome_fantasia="Outra Empresa Categorias",
            cnpj="00.000.000/0001-07",
        )
        conta_externa = ContaDRE.objects.create(
            empresa=outra_empresa,
            nome="Conta externa",
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            self.url,
            {
                f"conta_dre_{self.categoria_filha.pk}": conta_externa.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Uma das contas DRE selecionadas é inválida.")
        self.categoria_filha.refresh_from_db()
        self.assertIsNone(self.categoria_filha.conta_dre)

    def test_usuario_comum_nao_acessa_categorias(self):
        self.client.force_login(self.cliente)

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 403)

    def test_exporta_categorias_com_contas_dre_para_selecao(self):
        self.categoria_filha.conta_dre = self.conta_dre
        self.categoria_filha.save(update_fields=["conta_dre"])
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:exportar_planilha_categorias",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        ws = workbook["Categorias"]
        self.assertEqual([ws["A1"].value, ws["B1"].value], ["Categoria", "Conta DRE"])
        self.assertEqual(
            ws["A2"].value,
            f"{self.categoria_pai.codigo} - {self.categoria_pai.descricao}",
        )
        self.assertEqual(ws["B3"].value, self.conta_dre.nome)
        self.assertTrue(ws.data_validations.dataValidation)

    def test_importa_associacoes_de_categorias_pela_planilha(self):
        self.client.force_login(self.administrador)
        planilha = arquivo_xlsx(
            "Categorias",
            ("Categoria", "Conta DRE"),
            (
                (
                    f"{self.categoria_pai.codigo} - {self.categoria_pai.descricao}",
                    "",
                ),
                (
                    f"{self.categoria_filha.codigo} - {self.categoria_filha.descricao}",
                    self.conta_dre.nome,
                ),
            ),
        )

        response = self.client.post(
            reverse(
                "dashboards:importar_planilha_categorias",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {"planilha": planilha},
        )

        self.assertRedirects(response, self.url)
        self.categoria_filha.refresh_from_db()
        self.assertEqual(self.categoria_filha.conta_dre, self.conta_dre)


class SincronizacaoClientesOmieTests(TestCase):
    def setUp(self):
        self.administrador = get_user_model().objects.create_user(
            username="admin_sync_omie",
            password="senha-segura",
            is_staff=True,
        )
        self.empresa = Empresa.objects.create(
            nome="Empresa Sync OMIE Ltda",
            nome_fantasia="Empresa Sync OMIE",
            cnpj="00.000.000/0001-04",
        )
        self.integracao = IntegracaoOmie(
            empresa=self.empresa,
            app_key="app-key",
        )
        self.integracao.definir_app_secret("app-secret")
        self.integracao.save()

    @patch("apps.empresas.views.iniciar_sincronizacao_omie")
    def test_endpoint_inicia_sincronizacao(self, iniciar_mock):
        self.client.force_login(self.administrador)
        response = self.client.post(
            reverse(
                "dashboards:sincronizar_clientes_omie",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertEqual(response.status_code, 202)
        sincronizacao = SincronizacaoOmie.objects.get(empresa=self.empresa)
        iniciar_mock.assert_called_once_with(sincronizacao.pk)
        self.assertEqual(response.json()["status"], SincronizacaoOmie.Status.PENDENTE)

    @patch("apps.empresas.views.iniciar_sincronizacao_omie")
    def test_nova_sincronizacao_substitui_pendente_obsoleta(self, iniciar_mock):
        obsoleta = SincronizacaoOmie.objects.create(
            empresa=self.empresa,
            mensagem="Sincronização abandonada",
        )
        SincronizacaoOmie.objects.filter(pk=obsoleta.pk).update(
            atualizada_em=timezone.now() - timedelta(hours=1)
        )
        self.client.force_login(self.administrador)

        response = self.client.post(
            reverse(
                "dashboards:sincronizar_clientes_omie",
                kwargs={"empresa_slug": self.empresa.slug},
            )
        )

        self.assertEqual(response.status_code, 202)
        obsoleta.refresh_from_db()
        self.assertEqual(obsoleta.status, SincronizacaoOmie.Status.ERRO)
        nova = SincronizacaoOmie.objects.exclude(pk=obsoleta.pk).get()
        iniciar_mock.assert_called_once_with(nova.pk)
        self.assertEqual(response.json()["id"], nova.pk)

    def test_polling_encerra_sincronizacao_obsoleta(self):
        obsoleta = SincronizacaoOmie.objects.create(
            empresa=self.empresa,
            status=SincronizacaoOmie.Status.EM_ANDAMENTO,
            mensagem="Processando",
        )
        SincronizacaoOmie.objects.filter(pk=obsoleta.pk).update(
            atualizada_em=timezone.now() - timedelta(hours=1)
        )
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:status_sincronizacao_omie",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "sincronizacao_id": obsoleta.pk,
                },
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["status"], SincronizacaoOmie.Status.ERRO)
        self.assertIn("mais de 30 minutos", response.json()["erro"])

    @patch("apps.empresas.views._encerrar_sincronizacoes_omie_obsoletas")
    def test_polling_de_sincronizacao_ativa_nao_escreve_no_banco(
        self,
        encerrar_mock,
    ):
        ativa = SincronizacaoOmie.objects.create(
            empresa=self.empresa,
            status=SincronizacaoOmie.Status.EM_ANDAMENTO,
            mensagem="Processando",
        )
        self.client.force_login(self.administrador)

        response = self.client.get(
            reverse(
                "dashboards:status_sincronizacao_omie",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "sincronizacao_id": ativa.pk,
                },
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json()["status"],
            SincronizacaoOmie.Status.EM_ANDAMENTO,
        )
        encerrar_mock.assert_not_called()

    @patch("apps.empresas.omie.urlopen")
    def test_consultas_novas_usam_os_contratos_da_omie(self, urlopen_mock):
        urlopen_mock.return_value.__enter__.return_value.read.return_value = (
            b'{"pagina": 1, "total_de_paginas": 1, "total_de_registros": 0}'
        )

        consultar_tipos_conta_corrente(self.integracao, 2)
        consultar_contas_correntes(self.integracao, 3)
        consultar_contas_pagar(self.integracao, 4)
        consultar_contas_receber(self.integracao, 5)
        consultar_lancamentos_conta_corrente(self.integracao, 6)

        requisicao_tipos = urlopen_mock.call_args_list[0].args[0]
        payload_tipos = json.loads(requisicao_tipos.data)
        self.assertTrue(requisicao_tipos.full_url.endswith("/geral/tipocc/"))
        self.assertEqual(payload_tipos["call"], "ListarTiposCC")
        self.assertEqual(
            payload_tipos["param"][0],
            {"pagina": 2, "registros_por_pagina": 50},
        )

        requisicao_contas = urlopen_mock.call_args_list[1].args[0]
        payload_contas = json.loads(requisicao_contas.data)
        self.assertEqual(
            requisicao_contas.full_url.split("?")[0].rstrip("/").split("/")[-1],
            "contacorrente",
        )
        self.assertEqual(payload_contas["call"], "ListarContasCorrentes")
        self.assertEqual(
            payload_contas["param"][0],
            {
                "pagina": 3,
                "registros_por_pagina": 100,
                "apenas_importado_api": "N",
            },
        )

        requisicao_contas_pagar = urlopen_mock.call_args_list[2].args[0]
        payload_contas_pagar = json.loads(requisicao_contas_pagar.data)
        self.assertTrue(
            requisicao_contas_pagar.full_url.endswith("/financas/contapagar/")
        )
        self.assertEqual(payload_contas_pagar["call"], "ListarContasPagar")
        self.assertEqual(
            payload_contas_pagar["param"][0],
            {
                "pagina": 4,
                "registros_por_pagina": 20,
                "apenas_importado_api": "N",
            },
        )

        requisicao_contas_receber = urlopen_mock.call_args_list[3].args[0]
        payload_contas_receber = json.loads(requisicao_contas_receber.data)
        self.assertTrue(
            requisicao_contas_receber.full_url.endswith(
                "/financas/contareceber/"
            )
        )
        self.assertEqual(payload_contas_receber["call"], "ListarContasReceber")
        self.assertEqual(
            payload_contas_receber["param"][0],
            {
                "pagina": 5,
                "registros_por_pagina": 20,
                "apenas_importado_api": "N",
            },
        )

        requisicao_lancamentos = urlopen_mock.call_args_list[4].args[0]
        payload_lancamentos = json.loads(requisicao_lancamentos.data)
        self.assertTrue(
            requisicao_lancamentos.full_url.endswith(
                "/financas/contacorrentelancamentos/"
            )
        )
        self.assertEqual(payload_lancamentos["call"], "ListarLancCC")
        self.assertEqual(
            payload_lancamentos["param"][0],
            {
                "nPagina": 6,
                "nRegPorPagina": 20,
            },
        )

    @patch("apps.empresas.omie.consultar_lancamentos_conta_corrente")
    @patch("apps.empresas.omie.consultar_contas_receber")
    @patch("apps.empresas.omie.consultar_contas_pagar")
    @patch("apps.empresas.omie.consultar_contas_correntes")
    @patch("apps.empresas.omie.consultar_tipos_conta_corrente")
    @patch("apps.empresas.omie.consultar_categorias")
    @patch("apps.empresas.omie.consultar_departamentos")
    @patch("apps.empresas.omie.consultar_projetos")
    @patch("apps.empresas.omie.consultar_clientes")
    def test_importa_dados_omie_incluindo_contas_correntes(
        self,
        consultar_clientes_mock,
        consultar_projetos_mock,
        consultar_departamentos_mock,
        consultar_categorias_mock,
        consultar_tipos_conta_corrente_mock,
        consultar_contas_correntes_mock,
        consultar_contas_pagar_mock,
        consultar_contas_receber_mock,
        consultar_lancamentos_conta_corrente_mock,
    ):
        consultar_clientes_mock.side_effect = [
            {
                "pagina": 1,
                "total_de_paginas": 2,
                "total_de_registros": 2,
                "clientes_cadastro": [
                    {
                        "codigo_cliente_omie": 101,
                        "razao_social": "Cliente Um Ltda",
                        "nome_fantasia": "Cliente Um",
                        "cnpj_cpf": "00.000.000/0001-11",
                        "pessoa_fisica": "N",
                        "inativo": "N",
                        "tags": [{"tag": "Cliente"}],
                        "dadosBancarios": {"agencia": "1234"},
                    }
                ],
            },
            {
                "pagina": 2,
                "total_de_paginas": 2,
                "total_de_registros": 2,
                "clientes_cadastro": [
                    {
                        "codigo_cliente_omie": 202,
                        "razao_social": "Fornecedor Dois Ltda",
                        "cnpj_cpf": "00.000.000/0001-22",
                        "tags": [{"tag": "Fornecedor"}],
                    }
                ],
            },
        ]
        consultar_projetos_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "cadastro": [
                {
                    "codigo": 303,
                    "codInt": "PROJ-303",
                    "nome": "Projeto 303",
                    "inativo": "N",
                    "info": {"data_alt": "01/04/2025"},
                }
            ],
        }
        consultar_departamentos_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "departamentos": [
                {
                    "codigo": "5476993662",
                    "descricao": "Hinfoluz",
                    "estrutura": "001.001.001",
                    "inativo": "N",
                    "nivel_totalizador": "N",
                }
            ],
        }
        consultar_categorias_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "categoria_cadastro": [
                {
                    "categoria_superior": "0",
                    "codigo": "0.01",
                    "codigo_dre": "",
                    "conta_despesa": "N",
                    "conta_inativa": "N",
                    "conta_receita": "N",
                    "dadosDRE": {},
                    "definida_pelo_usuario": "N",
                    "descricao": "Transferência",
                    "descricao_padrao": "Transferência",
                    "id_conta_contabil": "",
                    "nao_exibir": "S",
                    "natureza": "",
                    "tag_conta_contabil": "",
                    "tipo_categoria": "",
                    "totalizadora": "S",
                    "transferencia": "S",
                }
            ],
        }
        consultar_tipos_conta_corrente_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "cadastros": [
                {
                    "cCodigo": "CX",
                    "cDescricao": "Caixinha",
                    "cGrupo": "CX",
                }
            ],
        }
        consultar_contas_correntes_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "ListarContasCorrentes": [
                {
                    "nCodCC": 3036783065,
                    "cCodCCInt": "CAIXA-01",
                    "tipo_conta_corrente": "CX",
                    "codigo_banco": "999",
                    "descricao": "Caixinha",
                    "codigo_agencia": "",
                    "numero_conta_corrente": "",
                    "saldo_inicial": 125.5,
                    "saldo_data": "08/10/2020",
                    "valor_limite": 500,
                    "nao_fluxo": "N",
                    "nao_resumo": "S",
                    "cobr_sn": "N",
                    "bol_sn": "N",
                    "pix_sn": "S",
                    "importado_api": "N",
                    "bloqueado": "N",
                    "inativo": "N",
                    "observacao": "Conta de caixa",
                }
            ],
        }
        consultar_contas_pagar_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "conta_pagar_cadastro": [
                {
                    "categorias": [
                        {
                            "codigo_categoria": "0.01",
                            "percentual": 100,
                            "valor": 74.89,
                        }
                    ],
                    "cnab_integracao_bancaria": {
                        "codigo_forma_pagamento": "PIX",
                        "pix_qrcode": "000201010212",
                    },
                    "codigo_categoria": "0.01",
                    "codigo_cliente_fornecedor": 202,
                    "codigo_lancamento_integracao": "",
                    "codigo_lancamento_omie": 1344090316,
                    "codigo_projeto": 303,
                    "codigo_tipo_documento": "BOL",
                    "data_emissao": "07/10/2025",
                    "data_entrada": "07/10/2025",
                    "data_previsao": "10/11/2025",
                    "data_vencimento": "04/11/2025",
                    "distribuicao": [],
                    "id_conta_corrente": 3036783065,
                    "id_origem": "APIP",
                    "info": {"cImpAPI": "N"},
                    "numero_documento": "RPS 3071844",
                    "numero_documento_fiscal": "3080816",
                    "numero_parcela": "001/001",
                    "retem_cofins": "N",
                    "retem_csll": "N",
                    "retem_inss": "N",
                    "retem_ir": "N",
                    "retem_iss": "N",
                    "retem_pis": "N",
                    "status_titulo": "ATRASADO",
                    "valor_documento": 74.89,
                }
            ],
        }
        consultar_contas_receber_mock.return_value = {
            "pagina": 1,
            "total_de_paginas": 1,
            "total_de_registros": 1,
            "conta_receber_cadastro": [
                {
                    "boleto": {
                        "cGerado": "N",
                        "cNumBancario": "",
                        "cNumBoleto": "",
                        "dDtEmBol": "",
                        "nPerJuros": 0,
                        "nPerMulta": 0,
                    },
                    "categorias": [
                        {
                            "codigo_categoria": "0.01",
                            "percentual": 100,
                            "valor": 500,
                        }
                    ],
                    "chave_nfe": "35201013835876000108550010000000841966893592",
                    "codigo_categoria": "0.01",
                    "codigo_cliente_fornecedor": 101,
                    "codigo_lancamento_integracao": "",
                    "codigo_lancamento_omie": 3037854759,
                    "codigo_projeto": 303,
                    "codigo_tipo_documento": "99999",
                    "data_emissao": "20/10/2020",
                    "data_previsao": "20/10/2020",
                    "data_registro": "20/10/2020",
                    "data_vencimento": "20/10/2020",
                    "distribuicao": [],
                    "id_conta_corrente": 3036783065,
                    "id_origem": "VENR",
                    "info": {"cImpAPI": "N"},
                    "nCodPedido": 3037850626,
                    "numero_documento_fiscal": "00000084",
                    "numero_parcela": "001/001",
                    "numero_pedido": "5",
                    "operacao": "11",
                    "retem_cofins": "N",
                    "retem_csll": "N",
                    "retem_inss": "N",
                    "retem_ir": "N",
                    "retem_iss": "N",
                    "retem_pis": "N",
                    "status_titulo": "CANCELADO",
                    "tipo_agrupamento": "I",
                    "valor_documento": 500,
                }
            ],
        }
        consultar_lancamentos_conta_corrente_mock.return_value = {
            "nPagina": 1,
            "nTotPaginas": 1,
            "nRegistros": 1,
            "nTotRegistros": 1,
            "listaLancamentos": [
                {
                    "cCodIntLanc": "",
                    "cabecalho": {
                        "dDtLanc": "08/04/2026",
                        "nCodCC": 3036783065,
                        "nValorLanc": 2289.5,
                    },
                    "departamentos": [
                        {
                            "cCodDep": "5476993662",
                            "nPerDep": 100,
                            "nValDep": 2289.5,
                        }
                    ],
                    "detalhes": {
                        "aCodCateg": [],
                        "cCodCateg": "0.01",
                        "cNumDoc": "",
                        "cObs": "",
                        "cTipo": "99999",
                        "nCodCliente": 101,
                        "nCodProjeto": 303,
                    },
                    "diversos": {
                        "cHrConc": "",
                        "cIdentLanc": "",
                        "cNatureza": "R",
                        "cOrigem": "BAXR",
                        "cUsConc": "",
                        "dDtConc": "",
                        "nCodComprador": 0,
                        "nCodLancCR": 3037854759,
                        "nCodVendedor": 0,
                    },
                    "info": {
                        "cImpAPI": "N",
                        "dAlt": "17/04/2026",
                        "dInc": "17/04/2026",
                    },
                    "nCodAgrup": 4353703652,
                    "nCodLanc": 4353703652,
                    "transferencia": {"nCodCCDestino": 0},
                }
            ],
        }
        sincronizacao = SincronizacaoOmie.objects.create(empresa=self.empresa)

        executar_sincronizacao_omie(sincronizacao.pk)

        sincronizacao.refresh_from_db()
        self.assertEqual(sincronizacao.status, SincronizacaoOmie.Status.CONCLUIDA)
        self.assertEqual(sincronizacao.pagina_atual, 10)
        self.assertEqual(sincronizacao.registros_processados, 10)
        self.assertEqual(CadastroOmie.objects.count(), 2)
        self.assertEqual(
            CadastroOmie.objects.get(codigo_cliente_omie=101).tipo,
            CadastroOmie.Tipo.CLIENTE,
        )
        self.assertEqual(
            CadastroOmie.objects.get(codigo_cliente_omie=202).tipo,
            CadastroOmie.Tipo.FORNECEDOR,
        )
        projeto = ProjetoOmie.objects.get(codigo=303)
        self.assertEqual(projeto.nome, "Projeto 303")
        self.assertEqual(projeto.codigo_integracao, "PROJ-303")
        departamento = DepartamentoOmie.objects.get(codigo="5476993662")
        self.assertEqual(departamento.descricao, "Hinfoluz")
        self.assertEqual(departamento.estrutura, "001.001.001")
        categoria = CategoriaOmie.objects.get(codigo="0.01")
        self.assertEqual(categoria.descricao, "Transferência")
        self.assertEqual(categoria.categoria_superior, "0")
        self.assertTrue(categoria.totalizadora)
        self.assertTrue(categoria.transferencia)
        self.assertTrue(categoria.nao_exibir)
        self.assertFalse(categoria.conta_inativa)
        tipo_conta = TipoContaCorrenteOmie.objects.get(codigo="CX")
        self.assertEqual(tipo_conta.descricao, "Caixinha")
        self.assertEqual(tipo_conta.grupo, "CX")
        conta = ContaCorrenteOmie.objects.get(codigo_omie=3036783065)
        self.assertEqual(conta.tipo_conta, tipo_conta)
        self.assertEqual(conta.tipo_codigo, "CX")
        self.assertEqual(conta.descricao, "Caixinha")
        self.assertEqual(str(conta.saldo_inicial), "125.50")
        self.assertEqual(str(conta.valor_limite), "500.00")
        self.assertTrue(conta.nao_resumo)
        self.assertTrue(conta.emite_pix)
        self.assertFalse(conta.inativo)
        self.assertEqual(conta.dados_originais["codigo_banco"], "999")
        conta_pagar = ContaPagarOmie.objects.get(
            codigo_lancamento_omie=1344090316
        )
        self.assertEqual(conta_pagar.fornecedor.codigo_cliente_omie, 202)
        self.assertEqual(conta_pagar.conta_corrente, conta)
        self.assertEqual(conta_pagar.categoria_principal, categoria)
        self.assertEqual(conta_pagar.projeto, projeto)
        self.assertEqual(conta_pagar.data_vencimento.isoformat(), "2025-11-04")
        self.assertEqual(str(conta_pagar.valor_documento), "74.89")
        self.assertEqual(str(conta_pagar.valor_a_pagar), "74.89")
        self.assertEqual(conta_pagar.status_titulo, "ATRASADO")
        self.assertEqual(
            conta_pagar.cnab_integracao_bancaria["codigo_forma_pagamento"],
            "PIX",
        )
        conta_receber = ContaReceberOmie.objects.get(
            codigo_lancamento_omie=3037854759
        )
        self.assertEqual(conta_receber.cliente.codigo_cliente_omie, 101)
        self.assertEqual(conta_receber.conta_corrente, conta)
        self.assertEqual(conta_receber.categoria_principal, categoria)
        self.assertEqual(conta_receber.projeto, projeto)
        self.assertEqual(
            conta_receber.data_vencimento.isoformat(),
            "2020-10-20",
        )
        self.assertEqual(str(conta_receber.valor_documento), "500.00")
        self.assertEqual(str(conta_receber.valor_a_receber), "500.00")
        self.assertEqual(conta_receber.status_titulo, "CANCELADO")
        self.assertEqual(conta_receber.boleto["cGerado"], "N")
        self.assertEqual(conta_receber.codigo_pedido_omie, 3037850626)
        self.assertEqual(conta_receber.tipo_agrupamento, "I")
        lancamento = LancamentoContaCorrenteOmie.objects.get(
            codigo_lancamento_omie=4353703652
        )
        self.assertEqual(lancamento.conta_corrente, conta)
        self.assertEqual(lancamento.categoria_principal, categoria)
        self.assertEqual(lancamento.cliente_fornecedor.codigo_cliente_omie, 101)
        self.assertEqual(lancamento.projeto, projeto)
        self.assertEqual(lancamento.conta_receber, conta_receber)
        self.assertEqual(lancamento.data_lancamento.isoformat(), "2026-04-08")
        self.assertEqual(str(lancamento.valor_lancamento), "2289.50")
        self.assertEqual(lancamento.natureza, "R")
        self.assertEqual(lancamento.origem, "BAXR")
        self.assertEqual(lancamento.departamentos[0]["cCodDep"], "5476993662")
        self.assertEqual(categoria.dados_originais["descricao"], "Transferência")
