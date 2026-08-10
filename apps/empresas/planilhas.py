from collections import Counter
from io import BytesIO
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .categorias import excluir_categorias_transferencia
from .models import CategoriaOmie, ContaDRE


class PlanilhaInvalida(ValueError):
    pass


AZUL = "2563C9"
AZUL_CLARO = "EAF1FF"


def _preparar_cabecalho(ws, titulos):
    for coluna, titulo in enumerate(titulos, start=1):
        celula = ws.cell(row=1, column=coluna, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=AZUL)
        celula.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(titulos))}1"
    ws.row_dimensions[1].height = 24


def _finalizar(workbook):
    arquivo = BytesIO()
    workbook.save(arquivo)
    arquivo.seek(0)
    return arquivo.getvalue()


def _abrir(arquivo, aba):
    try:
        workbook = load_workbook(arquivo, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError, KeyError) as exc:
        raise PlanilhaInvalida("O arquivo enviado não é uma planilha XLSX válida.") from exc
    if aba not in workbook.sheetnames:
        raise PlanilhaInvalida(f'A planilha precisa conter a aba "{aba}".')
    return workbook[aba]


def exportar_dre(empresa):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "DRE"
    _preparar_cabecalho(ws, ("Nome da conta DRE", "Tipo", "Operação"))
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    linha = 2
    contas_pai = (
        ContaDRE.objects.filter(empresa=empresa, conta_pai__isnull=True)
        .prefetch_related("contas_filhas")
        .order_by("ordem", "nome")
    )
    for conta in contas_pai:
        tipo_conta = "Resultado" if conta.eh_resultado else "Pai"
        ws.append((conta.nome, tipo_conta, conta.sinal))
        linha += 1
        if conta.eh_resultado:
            continue
        for filha in conta.contas_filhas.all():
            ws.append((filha.nome, "Filho", filha.sinal))
            linha += 1

    listas = workbook.create_sheet("Listas")
    listas.sheet_state = "hidden"
    listas.append(("Tipo", "Operação"))
    listas.append(("Pai", "+"))
    listas.append(("Filho", "-"))
    listas.append(("Resultado", "="))

    tipo = DataValidation(type="list", formula1="'Listas'!$A$2:$A$4")
    operacao = DataValidation(type="list", formula1="'Listas'!$B$2:$B$4")
    tipo.error = "Selecione Pai, Filho ou Resultado."
    operacao.error = "Selecione +, - ou =."
    tipo.errorTitle = operacao.errorTitle = "Valor inválido"
    tipo.showErrorMessage = operacao.showErrorMessage = True
    ws.add_data_validation(tipo)
    ws.add_data_validation(operacao)
    limite = max(linha + 200, 202)
    tipo.add(f"B2:B{limite}")
    operacao.add(f"C2:C{limite}")

    instrucoes = workbook.create_sheet("Instruções")
    instrucoes["A1"] = "Como preencher"
    instrucoes["A1"].font = Font(bold=True, color="FFFFFF")
    instrucoes["A1"].fill = PatternFill("solid", fgColor=AZUL)
    instrucoes["A2"] = "Cada conta Filho pertence à última conta Pai informada acima dela."
    instrucoes["A3"] = (
        "Contas Resultado ficam no nivel principal, usam Operação = e não recebem filhos."
    )
    instrucoes["A4"] = "Mantenha as linhas na ordem em que devem aparecer no DRE."
    instrucoes["A5"] = "Não altere os nomes das colunas nem o nome da aba DRE."
    instrucoes.column_dimensions["A"].width = 80
    return _finalizar(workbook)


def importar_dre(arquivo):
    ws = _abrir(arquivo, "DRE")
    cabecalho = tuple(ws.cell(1, coluna).value for coluna in range(1, 4))
    esperado = ("Nome da conta DRE", "Tipo", "Operação")
    if cabecalho != esperado:
        raise PlanilhaInvalida(
            "As colunas da aba DRE devem ser: Nome da conta DRE, Tipo e Operação."
        )

    contas = []
    pai_atual = None
    ordem_pais = 0
    ordem_filhas = 0
    for numero, valores in enumerate(
        ws.iter_rows(min_row=2, max_col=3, values_only=True), start=2
    ):
        nome, tipo, sinal = valores
        if all(valor in (None, "") for valor in valores):
            continue
        nome = str(nome or "").strip()
        tipo = str(tipo or "").strip().casefold()
        sinal = str(sinal or "").strip()
        if not nome:
            raise PlanilhaInvalida(f"Informe o nome da conta na linha {numero}.")
        if tipo not in {"pai", "filho", "resultado"}:
            raise PlanilhaInvalida(
                f'A coluna Tipo da linha {numero} deve conter "Pai", '
                '"Filho" ou "Resultado".'
            )
        if sinal not in {"+", "-", "="}:
            raise PlanilhaInvalida(
                f"A Operação da linha {numero} deve ser +, - ou =."
            )
        if sinal == "=" and tipo in {"pai", "resultado"}:
            tipo = "resultado"
        if tipo == "resultado" and sinal != "=":
            raise PlanilhaInvalida(
                f"A conta Resultado da linha {numero} deve usar Operação =."
            )
        if tipo == "filho" and sinal == "=":
            raise PlanilhaInvalida(
                f"A conta Resultado da linha {numero} não pode ser do tipo Filho."
            )
        if tipo in {"pai", "resultado"}:
            ordem_pais += 1
            ordem_filhas = 0
            pai_atual = None if tipo == "resultado" else len(contas)
            ordem = ordem_pais
        else:
            if pai_atual is None:
                raise PlanilhaInvalida(
                    f"A conta Filho da linha {numero} precisa estar abaixo de uma conta Pai."
                )
            ordem_filhas += 1
            ordem = ordem_filhas
        contas.append(
            {
                "nome": nome,
                "tipo": tipo,
                "sinal": sinal,
                "ordem": ordem,
                "indice_pai": pai_atual if tipo == "filho" else None,
            }
        )
    if not contas:
        raise PlanilhaInvalida("A planilha DRE não possui nenhuma conta preenchida.")
    return contas


def _rotulos_contas(empresa):
    contas = list(
        ContaDRE.objects.filter(empresa=empresa)
        .exclude(sinal=ContaDRE.Sinal.RESULTADO)
        .select_related("conta_pai")
        .order_by("conta_pai_id", "ordem", "nome")
    )
    rotulos_base = [
        conta.nome
        if conta.conta_pai_id is None
        else f"{conta.conta_pai.nome} > {conta.nome}"
        for conta in contas
    ]
    repeticoes = Counter(rotulos_base)
    rotulos = {
        conta.pk: (
            f"{rotulo} (ID {conta.pk})" if repeticoes[rotulo] > 1 else rotulo
        )
        for conta, rotulo in zip(contas, rotulos_base)
    }
    return contas, rotulos


def exportar_categorias(empresa):
    workbook = Workbook()
    ws = workbook.active
    ws.title = "Categorias"
    _preparar_cabecalho(ws, ("Categoria", "Conta DRE"))
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 48

    categorias = list(
        excluir_categorias_transferencia(
            CategoriaOmie.objects.filter(empresa=empresa, conta_inativa=False)
            .filter(ativo_omie=True)
        )
        .select_related("conta_dre")
        .order_by("codigo")
    )
    contas, rotulos = _rotulos_contas(empresa)
    for categoria in categorias:
        rotulo_categoria = f"{categoria.codigo} - {categoria.descricao}"
        rotulo_conta = (
            rotulos.get(categoria.conta_dre_id, "")
            if categoria.permite_vinculo_dre
            else ""
        )
        ws.append((rotulo_categoria, rotulo_conta))
        if not categoria.permite_vinculo_dre:
            ws.cell(ws.max_row, 2).fill = PatternFill("solid", fgColor="F0F2F5")

    listas = workbook.create_sheet("Contas DRE")
    listas.sheet_state = "hidden"
    listas["A1"] = "Conta DRE"
    for linha, conta in enumerate(contas, start=2):
        listas.cell(linha, 1, rotulos[conta.pk])

    if contas and categorias:
        validacao = DataValidation(
            type="list",
            formula1=f"'Contas DRE'!$A$2:$A${len(contas) + 1}",
            allow_blank=True,
        )
        validacao.error = "Selecione uma conta DRE disponível na lista."
        validacao.errorTitle = "Conta DRE inválida"
        validacao.showErrorMessage = True
        ws.add_data_validation(validacao)
        for linha, categoria in enumerate(categorias, start=2):
            if categoria.permite_vinculo_dre:
                validacao.add(ws.cell(linha, 2))

    instrucoes = workbook.create_sheet("Instruções")
    instrucoes["A1"] = "Como preencher"
    instrucoes["A1"].font = Font(bold=True, color="FFFFFF")
    instrucoes["A1"].fill = PatternFill("solid", fgColor=AZUL)
    instrucoes["A2"] = "Selecione na segunda coluna a conta DRE correspondente."
    instrucoes["A3"] = "Deixe a Conta DRE vazia para remover uma associação existente."
    instrucoes["A4"] = "Não altere a coluna Categoria nem o nome da aba Categorias."
    instrucoes.column_dimensions["A"].width = 80
    return _finalizar(workbook)


def importar_categorias(arquivo, empresa):
    ws = _abrir(arquivo, "Categorias")
    cabecalho = tuple(ws.cell(1, coluna).value for coluna in range(1, 3))
    if cabecalho != ("Categoria", "Conta DRE"):
        raise PlanilhaInvalida(
            "As colunas da aba Categorias devem ser: Categoria e Conta DRE."
        )

    categorias = list(
        excluir_categorias_transferencia(
            CategoriaOmie.objects.filter(empresa=empresa, conta_inativa=False)
            .filter(ativo_omie=True)
        )
        .order_by("codigo")
    )
    por_rotulo = {
        f"{categoria.codigo} - {categoria.descricao}": categoria
        for categoria in categorias
    }
    contas, rotulos = _rotulos_contas(empresa)
    contas_por_rotulo = {rotulos[conta.pk]: conta.pk for conta in contas}
    alteracoes = {}
    for numero, valores in enumerate(
        ws.iter_rows(min_row=2, max_col=2, values_only=True), start=2
    ):
        rotulo_categoria, rotulo_conta = valores
        if rotulo_categoria in (None, "") and rotulo_conta in (None, ""):
            continue
        rotulo_categoria = str(rotulo_categoria or "").strip()
        categoria = por_rotulo.get(rotulo_categoria)
        if categoria is None:
            raise PlanilhaInvalida(
                f"A categoria da linha {numero} não pertence à base ativa desta empresa."
            )
        rotulo_conta = str(rotulo_conta or "").strip()
        if not categoria.permite_vinculo_dre:
            if rotulo_conta:
                raise PlanilhaInvalida(
                    f"A categoria da linha {numero} não permite associação ao DRE."
                )
            continue
        if rotulo_conta and rotulo_conta not in contas_por_rotulo:
            raise PlanilhaInvalida(
                f'A Conta DRE "{rotulo_conta}" da linha {numero} não é válida.'
            )
        alteracoes[categoria.pk] = contas_por_rotulo.get(rotulo_conta)
    if not alteracoes:
        raise PlanilhaInvalida(
            "A planilha não possui categorias disponíveis para importação."
        )
    return alteracoes
