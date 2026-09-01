import json
from io import BytesIO
from datetime import date, timedelta
from decimal import Decimal
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from apps.dashboards.dre_services import dre_gerencial
from apps.dashboards.faturamento_services import faturamento_comercial
from apps.dashboards.fluxo_caixa_services import fluxo_de_caixa
from apps.dashboards.margem_rentabilidade_services import margem_rentabilidade_comercial
from apps.dashboards.models import AprovacaoPagamento
from apps.dashboards.visao_geral_services import visao_geral_financeira
from apps.empresas.models import (
    CadastroOmie,
    CategoriaOmie,
    ContaCorrenteOmie,
    ContaDRE,
    ContaPagarOmie,
    ContaReceberOmie,
    DepartamentoOmie,
    Empresa,
    EmpresaUsuario,
    IntegracaoOmie,
    LancamentoContaCorrenteOmie,
    MetaVendedorComercial,
    MovimentoFinanceiroOmie,
    NfseOmie,
    OrdemServicoItemOmie,
    OrdemServicoOmie,
    PedidoCompraItemOmie,
    PedidoCompraOmie,
    PedidoItemOmie,
    PedidoOmie,
    PesqTituloFinanceiroOmie,
    PosicaoEstoqueOmie,
    ProdutoOmie,
    ProjetoOmie,
    RecebimentoNfeItemOmie,
    RecebimentoNfeOmie,
    ServicoOmie,
    SincronizacaoOmie,
    VendedorOmie,
)
from apps.compras.models import BudgetConfiguracaoCompra, BudgetLimiteCompra


class DashboardPermissaoTests(TestCase):
    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="analista", password="senha-segura"
        )
        self.empresa = Empresa.objects.create(
            nome="Oeste Cliente Ltda",
            nome_fantasia="Oeste Cliente",
            cnpj="11.111.111/0001-11",
            grupo="Grupo Oeste",
        )

    def test_usuario_sem_vinculo_nao_acessa_dashboard(self):
        self.client.force_login(self.usuario)
        response = self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug})
        )
        self.assertEqual(response.status_code, 404)

    def test_usuario_vinculado_acessa_todas_as_areas_padrao(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        for area in ("comercial", "financeiro", "compras", "estoque", "crm"):
            response = self.client.get(
                reverse(
                    "dashboards:area",
                    kwargs={"empresa_slug": self.empresa.slug, "area_slug": area},
                )
            )
            self.assertEqual(response.status_code, 200)

    def test_menu_principal_usa_imagens_das_areas(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug})
        )

        for imagem in (
            "vendas.png",
            "financeiro.png",
            "compras.png",
            "estoque.png",
            "crm.png",
        ):
            self.assertContains(response, f"/media/{imagem}")

    def test_parametros_aparecem_acima_do_status_para_administrador(self):
        administrador = get_user_model().objects.create_user(
            username="admin_dashboard",
            password="senha-segura",
            is_staff=True,
        )
        self.client.force_login(administrador)

        response = self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug})
        )

        url_parametros = reverse(
            "dashboards:parametros",
            kwargs={"empresa_slug": self.empresa.slug},
        )
        self.assertContains(response, url_parametros)
        self.assertNotContains(response, "Estrutura pronta para integra")

    def test_parametros_nao_aparecem_para_usuario_comum(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug})
        )

        self.assertNotContains(
            response,
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
        )

    def test_administrador_da_empresa_ve_parametros_sem_ser_staff(self):
        EmpresaUsuario.objects.create(
            empresa=self.empresa,
            usuario=self.usuario,
            papel=EmpresaUsuario.Papel.ADMINISTRADOR,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug})
        )

        self.assertContains(
            response,
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
        )

    def test_gestor_da_empresa_ve_parametros_na_home(self):
        EmpresaUsuario.objects.create(
            empresa=self.empresa,
            usuario=self.usuario,
            papel=EmpresaUsuario.Papel.GESTOR,
            areas_permitidas=["financeiro"],
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug})
        )

        self.assertContains(
            response,
            reverse(
                "dashboards:parametros",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
        )

    def test_gestor_acessa_apenas_modulos_liberados(self):
        EmpresaUsuario.objects.create(
            empresa=self.empresa,
            usuario=self.usuario,
            papel=EmpresaUsuario.Papel.GESTOR,
            areas_permitidas=["financeiro"],
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug})
        )
        self.assertContains(response, "Financeiro")
        self.assertNotContains(response, "Comercial")
        self.assertEqual(
            self.client.get(
                reverse(
                    "dashboards:area",
                    kwargs={
                        "empresa_slug": self.empresa.slug,
                        "area_slug": "comercial",
                    },
                )
            ).status_code,
            404,
        )

    def test_usuario_visualiza_apenas_dashboard_liberado(self):
        EmpresaUsuario.objects.create(
            empresa=self.empresa,
            usuario=self.usuario,
            papel=EmpresaUsuario.Papel.VISUALIZADOR,
            areas_permitidas=["comercial"],
            dashboards_permitidos=["comercial:faturamento"],
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:area",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                },
            )
        )
        self.assertContains(response, "Faturamento")
        self.assertNotContains(response, "Desempenho de vendedores")
        self.assertEqual(
            self.client.get(
                reverse(
                    "dashboards:dashboard",
                    kwargs={
                        "empresa_slug": self.empresa.slug,
                        "area_slug": "comercial",
                        "dashboard_slug": "desempenho-de-vendedores",
                    },
                )
            ).status_code,
            404,
        )

    def test_area_inexistente_retorna_404(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)
        response = self.client.get(
            reverse(
                "dashboards:area",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "inexistente",
                },
            )
        )
        self.assertEqual(response.status_code, 404)

    def test_area_exibe_os_dashboards_disponiveis(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:area",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                },
            )
        )

        self.assertContains(response, "Faturamento")
        self.assertContains(response, "Desempenho de vendedores")
        self.assertContains(response, "Análise de clientes")
        self.assertContains(response, "Margem e Rentabilidade")

    def test_area_exibe_data_da_ultima_sincronizacao_concluida(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        finalizada_em = timezone.now()
        SincronizacaoOmie.objects.create(
            empresa=self.empresa,
            status=SincronizacaoOmie.Status.CONCLUIDA,
            finalizada_em=finalizada_em,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:area",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                },
            )
        )

        self.assertContains(response, "Atualizado -")
        self.assertContains(
            response,
            timezone.localtime(finalizada_em).strftime("%d/%m/%Y %H:%M"),
        )

    def test_dashboard_da_area_pode_ser_aberto(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "faturamento",
                },
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Faturamento")

    def test_filtro_de_categorias_aparece_apenas_no_financeiro(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        categoria_pai = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01",
            descricao="Receitas",
            totalizadora=True,
        )
        CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.01",
            categoria_superior=categoria_pai.codigo,
            descricao="Receita de servicos",
        )
        self.client.force_login(self.usuario)

        financeiro = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                    "dashboard_slug": "visao-geral",
                },
            )
        )
        comercial = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "faturamento",
                },
            )
        )

        self.assertContains(financeiro, "Categorias")
        self.assertContains(financeiro, "Receitas")
        self.assertContains(financeiro, "Receita de servicos")
        self.assertNotContains(comercial, "Categorias")

    def test_area_compras_exibe_dashboard_budget(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:area",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "compras",
                },
            )
        )

        self.assertContains(response, "Budget")
        self.assertContains(response, "Compare budget")
        self.assertContains(response, "Score de fornecedor")
        self.assertContains(response, "Curva ABC")
        self.assertContains(response, "Análise de Preço")
        self.assertContains(response, "Saving")
        self.assertContains(response, "analise-preco-saving")
        self.assertNotContains(response, "evolucao-de-compras")
        self.assertNotContains(response, "analise-de-fornecedores")
        self.assertNotContains(response, "pedidos-de-compra")
        self.assertNotContains(response, "prazos-de-entrega")

    def test_dashboard_budget_calcula_consumo_por_produto(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        configuracao = BudgetConfiguracaoCompra.objects.create(
            empresa=self.empresa,
            tipos_controle=[BudgetConfiguracaoCompra.TipoControle.PRODUTO],
        )
        produto = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=100,
            codigo="PRD-100",
            descricao="Acionador Manual Enderecavel AME 566",
        )
        BudgetLimiteCompra.objects.create(
            empresa=self.empresa,
            configuracao=configuracao,
            tipo_controle=BudgetConfiguracaoCompra.TipoControle.PRODUTO,
            referencia_codigo=str(produto.codigo_produto),
            referencia_nome=produto.descricao,
            limite_compra=Decimal("1000"),
        )
        pedido = PedidoCompraOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=500,
            numero_pedido="PC-500",
            data_previsao=date(ano_atual, 1, 15),
        )
        PedidoCompraItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido,
            codigo_item=501,
            produto=produto,
            codigo_produto=produto.codigo_produto,
            descricao=produto.descricao,
            quantidade=1,
            valor_unitario=1200,
            valor_total=1200,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "compras",
                    "dashboard_slug": "budget",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
                "budget_dimensao": "produto",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Orcamento total")
        self.assertContains(response, "Gasto realizado")
        self.assertContains(response, "Itens estourados")
        self.assertContains(response, "Acionador Manual Enderecavel AME 566")
        self.assertContains(response, "120%")
        self.assertContains(response, "Dimensao: Produto")
        self.assertContains(response, 'name="budget_dimensao"')
        self.assertNotContains(response, "Todos os projetos")
        self.assertNotContains(response, "Todos os departamentos")
        self.assertEqual(response.context["budget"]["itens_estourados"], 1)

    def test_dashboard_budget_mensal_considera_meses_selecionados(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        configuracao = BudgetConfiguracaoCompra.objects.create(
            empresa=self.empresa,
            tipos_controle=[BudgetConfiguracaoCompra.TipoControle.PRODUTO],
            periodicidades_controle={
                BudgetConfiguracaoCompra.TipoControle.PRODUTO: "mensal"
            },
            meses_controle={BudgetConfiguracaoCompra.TipoControle.PRODUTO: [6]},
        )
        produto = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=100,
            codigo="PRD-100",
            descricao="Produto mensal",
        )
        BudgetLimiteCompra.objects.create(
            empresa=self.empresa,
            configuracao=configuracao,
            tipo_controle=BudgetConfiguracaoCompra.TipoControle.PRODUTO,
            referencia_codigo=str(produto.codigo_produto),
            referencia_nome=produto.descricao,
            limite_compra=Decimal("1000"),
        )
        pedido = PedidoCompraOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=510,
            numero_pedido="PC-510",
            data_previsao=date(ano_atual, 6, 10),
        )
        PedidoCompraItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido,
            codigo_item=511,
            produto=produto,
            codigo_produto=produto.codigo_produto,
            descricao=produto.descricao,
            valor_total=Decimal("2000"),
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "compras",
                    "dashboard_slug": "budget",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"ano-{ano_atual}",
                "budget_dimensao": "produto",
            },
        )

        self.assertEqual(response.context["budget"]["total_budget"], Decimal("1000"))
        self.assertEqual(response.context["budget"]["itens_estourados"], 1)

    def test_dashboard_budget_calcula_consumo_por_departamento(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        departamento = DepartamentoOmie.objects.create(
            empresa=self.empresa,
            codigo="DEP-001",
            descricao="Administrativo",
            estrutura="001.001",
        )
        configuracao = BudgetConfiguracaoCompra.objects.create(
            empresa=self.empresa,
            tipos_controle=[BudgetConfiguracaoCompra.TipoControle.DEPARTAMENTO],
        )
        BudgetLimiteCompra.objects.create(
            empresa=self.empresa,
            configuracao=configuracao,
            tipo_controle=BudgetConfiguracaoCompra.TipoControle.DEPARTAMENTO,
            referencia_codigo=departamento.codigo,
            referencia_nome=departamento.descricao,
            limite_compra=Decimal("1000"),
        )
        pedido = PedidoCompraOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=520,
            numero_pedido="PC-520",
            data_previsao=date(ano_atual, 2, 10),
            departamentos_consulta=[
                {"cCodDep": departamento.codigo, "cDesDep": departamento.descricao}
            ],
        )
        PedidoCompraItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido,
            codigo_item=521,
            descricao="Compra administrativa",
            valor_total=Decimal("850"),
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "compras",
                    "dashboard_slug": "budget",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-02",
                "budget_dimensao": "departamento",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Dimensao: Departamento")
        self.assertContains(response, "Administrativo")
        self.assertContains(response, "85%")
        self.assertEqual(response.context["budget"]["status"]["atencao"], 1)

    def test_dashboard_score_fornecedor_exibe_estrutura_inicial(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        fornecedor_a = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=101,
            razao_social="Fornecedor Alpha",
            tipo="fornecedor",
        )
        fornecedor_b = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=202,
            razao_social="Fornecedor Beta",
            tipo="fornecedor",
        )
        produto_a = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=1001,
            codigo="PRD-A",
            descricao="Produto A",
        )
        produto_b = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=1002,
            codigo="PRD-B",
            descricao="Produto B",
        )
        pedido_anterior = PedidoCompraOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=10,
            numero_pedido="10",
            etapa="10",
            codigo_fornecedor=101,
            fornecedor=fornecedor_a,
            data_inclusao=date(2026, 7, 1),
            data_previsao=date(2026, 7, 5),
        )
        PedidoCompraItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido_anterior,
            codigo_item=10,
            produto=produto_a,
            codigo_produto=produto_a.codigo_produto,
            codigo_produto_texto=produto_a.codigo,
            descricao=produto_a.descricao,
            quantidade=10,
            quantidade_recebida=10,
            valor_unitario=100,
            valor_total=1000,
        )
        pedido_a = PedidoCompraOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=11,
            numero_pedido="119",
            etapa="15",
            codigo_fornecedor=101,
            fornecedor=fornecedor_a,
            data_inclusao=date(2026, 8, 4),
            data_previsao=date(2026, 8, 10),
        )
        PedidoCompraItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido_a,
            codigo_item=11,
            produto=produto_a,
            codigo_produto=produto_a.codigo_produto,
            codigo_produto_texto=produto_a.codigo,
            descricao=produto_a.descricao,
            quantidade=10,
            quantidade_recebida=10,
            valor_unitario=92,
            valor_total=920,
        )
        RecebimentoNfeOmie.objects.create(
            empresa=self.empresa,
            codigo_recebimento=100,
            data_registro=date(2026, 8, 10),
            numero_nfe="100",
        )
        RecebimentoNfeItemOmie.objects.create(
            empresa=self.empresa,
            recebimento=RecebimentoNfeOmie.objects.get(codigo_recebimento=100),
            codigo_recebimento=100,
            sequencia=1,
            numero_pedido_compra="119",
            data_recebimento=date(2026, 8, 10),
            codigo_produto_texto=produto_a.codigo,
            quantidade_nfe=10,
            quantidade_recebida=10,
            preco_unitario=92,
        )
        pedido_b = PedidoCompraOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=12,
            numero_pedido="120",
            etapa="10",
            codigo_fornecedor=202,
            fornecedor=fornecedor_b,
            data_inclusao=date(2026, 8, 5),
            data_previsao=date(2026, 8, 10),
        )
        PedidoCompraItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido_b,
            codigo_item=12,
            produto=produto_b,
            codigo_produto=produto_b.codigo_produto,
            codigo_produto_texto=produto_b.codigo,
            descricao=produto_b.descricao,
            quantidade=10,
            quantidade_recebida=5,
            valor_unitario=50,
            valor_total=500,
        )
        RecebimentoNfeOmie.objects.create(
            empresa=self.empresa,
            codigo_recebimento=101,
            data_registro=date(2026, 8, 12),
            numero_nfe="101",
        )
        RecebimentoNfeItemOmie.objects.create(
            empresa=self.empresa,
            recebimento=RecebimentoNfeOmie.objects.get(codigo_recebimento=101),
            codigo_recebimento=101,
            sequencia=1,
            numero_pedido_compra="120",
            data_recebimento=date(2026, 8, 12),
            codigo_produto_texto=produto_b.codigo,
            quantidade_nfe=5,
            quantidade_recebida=5,
            preco_unitario=50,
        )
        pedido_ignorado = PedidoCompraOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=13,
            numero_pedido="121",
            etapa="20",
            codigo_fornecedor=202,
            fornecedor=fornecedor_b,
            data_inclusao=date(2026, 8, 6),
            data_previsao=date(2026, 8, 10),
        )
        PedidoCompraItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido_ignorado,
            codigo_item=13,
            produto=produto_b,
            codigo_produto=produto_b.codigo_produto,
            codigo_produto_texto=produto_b.codigo,
            descricao=produto_b.descricao,
            quantidade=10,
            quantidade_recebida=10,
            valor_unitario=1,
            valor_total=10,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "compras",
                    "dashboard_slug": "score-de-fornecedor",
                },
            ),
            {"_filtrar": "1", "periodo": "mes-2026-08"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Score medio da base")
        self.assertContains(response, "OTD medio geral")
        self.assertContains(response, "A-excelente")
        self.assertContains(response, "B - Muito bom")
        self.assertContains(response, "C - Bom")
        self.assertContains(response, "D- Atencao")
        self.assertContains(response, "3 principais fornecedores do grupo", count=4)
        self.assertContains(response, "Ranking geral")
        self.assertContains(response, "10 melhores scores")
        self.assertContains(response, "Fornecedor Alpha")
        self.assertContains(response, "Fornecedor Beta")
        self.assertIn("score_fornecedor", response.context)
        dados = response.context["score_fornecedor"]
        self.assertEqual(len(dados["ranking"]), 2)
        self.assertEqual(dados["ranking"][0]["nome"], "Fornecedor Alpha")
        self.assertEqual(dados["ranking"][0]["otd"], 100)
        self.assertEqual(dados["ranking"][0]["conformidade_nf"], 100)
        self.assertEqual(dados["ranking"][0]["estabilidade_preco"], 108)
        self.assertEqual(dados["ranking"][1]["otd"], 60)
        self.assertEqual(dados["ranking"][1]["conformidade_nf"], 50)
        self.assertEqual(dados["ranking"][1]["itens"], 1)

    def test_dashboard_curva_abc_fornecedores_exibe_layout_inicial(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        fornecedor_alfa = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=301,
            razao_social="Fornecedor Alfa",
            tipo="fornecedor",
        )
        fornecedor_beta = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=302,
            razao_social="Fornecedor Beta",
            tipo="fornecedor",
        )
        fornecedor_gama = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=303,
            razao_social="Fornecedor Gama",
            tipo="fornecedor",
        )
        pedidos = [
            (701, "701", fornecedor_alfa, Decimal("600")),
            (702, "702", fornecedor_beta, Decimal("300")),
            (703, "703", fornecedor_gama, Decimal("100")),
        ]
        for codigo_pedido, numero_pedido, fornecedor, valor in pedidos:
            pedido = PedidoCompraOmie.objects.create(
                empresa=self.empresa,
                codigo_pedido=codigo_pedido,
                numero_pedido=numero_pedido,
                etapa="10",
                codigo_fornecedor=fornecedor.codigo_cliente_omie,
                fornecedor=fornecedor,
                data_previsao=date(2026, 8, 10),
            )
            PedidoCompraItemOmie.objects.create(
                empresa=self.empresa,
                pedido=pedido,
                codigo_item=codigo_pedido,
                descricao=f"Compra {numero_pedido}",
                quantidade=1,
                valor_unitario=valor,
                valor_total=valor,
            )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "compras",
                    "dashboard_slug": "curva-abc",
                },
            ),
            {"_filtrar": "1", "periodo": "mes-2026-08"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Gasto total")
        self.assertContains(response, "Fornecedores ativos")
        self.assertContains(response, "Concentracao top 10")
        self.assertContains(response, "Fornecedores classe A")
        self.assertContains(response, "Gasto por fornecedor e % acumulado")
        self.assertContains(response, "Distribuicao por classe")
        self.assertContains(response, "Ranking de fornecedores")
        self.assertContains(response, "Fornecedor Alfa")
        self.assertContains(response, "60.0%")
        self.assertContains(response, "90.0%")
        self.assertContains(response, "100.0%")
        self.assertIn("curva_abc_fornecedores", response.context)
        dados = response.context["curva_abc_fornecedores"]
        self.assertEqual(dados["fornecedores_ativos"], 3)
        self.assertEqual(dados["gasto_total"], "R$ 1 mil")
        self.assertEqual(dados["concentracao_top10"], "100%")
        self.assertEqual(dados["fornecedores"][0]["nome"], "Fornecedor Alfa")
        self.assertEqual(dados["fornecedores"][1]["nome"], "Fornecedor Beta")
        self.assertEqual(dados["fornecedores"][2]["nome"], "Fornecedor Gama")
        self.assertEqual(dados["fornecedores"][0]["classe"], "A")
        self.assertEqual(dados["fornecedores"][1]["classe"], "B")
        self.assertEqual(dados["fornecedores"][2]["classe"], "C")
        self.assertTrue(dados["curva_pontos"])

    def test_dashboard_analise_preco_saving_usa_recebimentos_e_cmc(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        produto_a = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=901,
            codigo="PRD901",
            descricao="Produto Saving",
            descricao_familia="Familia A",
        )
        produto_b = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=902,
            codigo="PRD902",
            descricao="Produto Alta",
            descricao_familia="Familia B",
        )
        produto_ignorado = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=903,
            codigo="PRD903",
            descricao="Produto Ignorado",
            descricao_familia="Familia C",
        )
        PosicaoEstoqueOmie.objects.create(
            empresa=self.empresa,
            produto=produto_a,
            codigo_produto=produto_a.codigo_produto,
            codigo_local_estoque=1,
            codigo=produto_a.codigo,
            descricao=produto_a.descricao,
            data_posicao=date(2026, 8, 1),
            saldo=100,
            cmc=98,
        )
        PosicaoEstoqueOmie.objects.create(
            empresa=self.empresa,
            produto=produto_b,
            codigo_produto=produto_b.codigo_produto,
            codigo_local_estoque=1,
            codigo=produto_b.codigo,
            descricao=produto_b.descricao,
            data_posicao=date(2026, 8, 1),
            saldo=100,
            cmc=51,
        )

        def criar_item(codigo_recebimento, produto, data_emissao, preco, etapa="60", ajustes=None):
            recebimento = RecebimentoNfeOmie.objects.create(
                empresa=self.empresa,
                codigo_recebimento=codigo_recebimento,
                etapa=etapa,
                data_emissao_nfe=data_emissao,
                codigo_fornecedor=700 + codigo_recebimento,
                cabec={"cNome": f"Fornecedor {codigo_recebimento}"},
            )
            return RecebimentoNfeItemOmie.objects.create(
                empresa=self.empresa,
                recebimento=recebimento,
                codigo_recebimento=codigo_recebimento,
                sequencia=1,
                data_recebimento=data_emissao,
                codigo_produto_texto=produto.codigo,
                descricao=produto.descricao,
                quantidade_nfe=10,
                quantidade_recebida=10,
                preco_unitario=preco,
                valor_total_item=Decimal(str(preco)) * Decimal("10"),
                item_ajustes=ajustes or {"cNaoGerarFinanceiro": "N"},
            )

        criar_item(801, produto_a, date(2026, 7, 1), 90)
        criar_item(802, produto_a, date(2026, 7, 20), 95)
        criar_item(803, produto_a, date(2026, 8, 10), 80)
        criar_item(804, produto_b, date(2026, 7, 5), 50)
        criar_item(805, produto_b, date(2026, 8, 12), 60)
        criar_item(806, produto_ignorado, date(2026, 8, 12), 10, etapa="40")
        criar_item(
            807,
            produto_ignorado,
            date(2026, 8, 13),
            10,
            ajustes={"cNaoGerarFinanceiro": "S"},
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "compras",
                    "dashboard_slug": "analise-preco-saving",
                },
            ),
            {"_filtrar": "1", "periodo": "mes-2026-08"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Impacto por familia de produto")
        self.assertContains(response, "Familia A")
        self.assertContains(response, "Produto Saving")
        self.assertContains(response, "Produto Alta")
        self.assertNotContains(response, "Produto Ignorado")
        self.assertIn("analise_preco_saving", response.context)
        dados = response.context["analise_preco_saving"]
        self.assertEqual(dados["saving_total"], "R$ 200,00")
        self.assertEqual(dados["perda_total"], "R$ 100,00")
        self.assertEqual(dados["itens_alta"], 1)
        self.assertEqual(dados["familia_destaque"], "Familia A")
        self.assertEqual(dados["itens"][0]["base_origem"], "CMC anterior")

    def test_dashboard_exibe_filtros_e_projetos_ativos(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        projeto = ProjetoOmie.objects.create(
            empresa=self.empresa,
            codigo=123,
            nome="PC2025-04",
        )
        ProjetoOmie.objects.create(
            empresa=self.empresa,
            codigo=456,
            nome="Projeto inativo",
            inativo=True,
        )
        departamento = DepartamentoOmie.objects.create(
            empresa=self.empresa,
            codigo="5476993662",
            descricao="Hinfoluz",
            estrutura="001.001.001",
        )
        DepartamentoOmie.objects.create(
            empresa=self.empresa,
            codigo="999",
            descricao="Departamento inativo",
            inativo=True,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "faturamento",
                },
            ),
            {
                "projeto": projeto.codigo,
                "departamento": departamento.codigo,
            },
        )

        self.assertContains(response, "Todos os projetos")
        self.assertContains(response, "PC2025-04")
        self.assertNotContains(response, "Projeto inativo")
        self.assertContains(response, "Todos os departamentos")
        self.assertContains(response, "Hinfoluz")
        self.assertNotContains(response, "Departamento inativo")
        self.assertContains(response, self.empresa.nome_fantasia)

    def test_faturamento_exibe_indicadores_graficos_top5_e_filtros(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        vendedor = VendedorOmie.objects.create(
            empresa=self.empresa,
            codigo=77,
            nome="Maria Vendas",
        )
        MetaVendedorComercial.objects.create(
            empresa=self.empresa,
            vendedor=vendedor,
            ano=ano_atual,
            mes=1,
            valor_mensal=10000,
        )
        produto = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=100,
            codigo="NB-15",
            descricao="Notebook Pro 15",
        )
        cliente = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=150,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente Excel",
        )
        servico = ServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_servico=200,
            codigo="SUP",
            descricao="Suporte tecnico",
        )
        pedido = PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=300,
            numero_pedido="PV-300",
            codigo_cliente=cliente.codigo_cliente_omie,
            cliente=cliente,
            codigo_vendedor=vendedor.codigo,
            data_inclusao=date(ano_atual, 1, 5),
            data_faturamento=date(ano_atual, 1, 8),
            faturado=True,
            valor_mercadorias=2700,
            valor_frete=100,
            valor_seguro=50,
            frete={"outras_despesas": 150},
            valor_total_pedido=3000,
            dados_originais={"nNF": "NF-9001"},
        )
        PedidoItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido,
            codigo_item=301,
            produto=produto,
            codigo_produto=produto.codigo_produto,
            descricao=produto.descricao,
            quantidade=2,
            valor_unitario=1500,
            valor_total=3000,
            valor_mercadoria=2700,
            imposto={"valor_icms": 200},
        )
        ordem = OrdemServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_os=400,
            numero_os="OS-400",
            codigo_vendedor=vendedor.codigo,
            codigo_cliente=cliente.codigo_cliente_omie,
            cliente=cliente,
            data_inclusao=date(ano_atual, 1, 7),
            data_faturamento=date(ano_atual, 1, 9),
            faturada=True,
            valor_total=1200,
            numero_recibo="123",
            dados_originais={
                "ListaRpsNfse": [
                    {
                        "nNfse": "8859",
                    }
                ],
            },
        )
        NfseOmie.objects.create(
            empresa=self.empresa,
            codigo_nf=9001,
            numero_nfse="8859",
            status_nfse="F",
            codigo_os=ordem.codigo_os,
            numero_os=ordem.numero_os,
            ordem_servico=ordem,
            data_emissao=ordem.data_faturamento,
            valor_nfse=1200,
        )
        ordem_rejeitada = OrdemServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_os=401,
            numero_os="OS-401",
            codigo_vendedor=vendedor.codigo,
            codigo_cliente=cliente.codigo_cliente_omie,
            cliente=cliente,
            data_inclusao=date(ano_atual, 1, 10),
            data_faturamento=date(ano_atual, 1, 11),
            faturada=True,
            valor_total=999,
        )
        NfseOmie.objects.create(
            empresa=self.empresa,
            codigo_nf=9002,
            numero_nfse="8860",
            status_nfse="N",
            codigo_os=ordem_rejeitada.codigo_os,
            numero_os=ordem_rejeitada.numero_os,
            ordem_servico=ordem_rejeitada,
            data_emissao=ordem_rejeitada.data_faturamento,
            valor_nfse=999,
        )
        OrdemServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_os=402,
            numero_os="OS-402",
            codigo_vendedor=vendedor.codigo,
            codigo_cliente=cliente.codigo_cliente_omie,
            cliente=cliente,
            data_inclusao=date(ano_atual, 1, 12),
            data_faturamento=date(ano_atual, 1, 13),
            faturada=True,
            valor_total=800,
            numero_recibo="456",
        )
        OrdemServicoItemOmie.objects.create(
            empresa=self.empresa,
            ordem_servico=ordem,
            codigo_item=401,
            servico=servico,
            codigo_servico=servico.codigo_servico,
            descricao=servico.descricao,
            quantidade=3,
            valor_unitario=400,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "faturamento",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
                "vendedor": f"{self.empresa.pk}:{vendedor.codigo}",
                "tipo_faturamento": ["produtos", "servicos"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Faturado")
        self.assertContains(response, "Meta do periodo")
        self.assertContains(response, "R$ 10.00 Mil")
        self.assertContains(response, "Pedidos emitidos")
        self.assertContains(response, "Ticket medio")
        self.assertContains(response, "Maria Vendas")
        self.assertContains(response, "Produtos")
        self.assertContains(response, "Servicos")
        self.assertContains(response, "Impostos")
        self.assertContains(response, "Notebook Pro 15")
        self.assertContains(response, "Suporte tecnico")
        self.assertContains(response, "data-billing-main-chart")
        self.assertContains(response, "data-billing-goal-chart")
        self.assertContains(response, "Excel Produtos")
        self.assertContains(response, "Excel Serviços")
        self.assertContains(response, "Exportar produtos para Excel")
        self.assertContains(response, "Exportar serviços para Excel")
        self.assertContains(response, "[10000.0]")
        self.assertEqual(
            response.context["faturamento"]["produtos_mercadorias"],
            [2700.0],
        )
        self.assertEqual(
            response.context["faturamento"]["produtos_frete_despesas"],
            [300.0],
        )
        self.assertEqual(
            response.context["faturamento"]["produtos_impostos"],
            [0],
        )
        self.assertEqual(
            response.context["faturamento"]["indicadores"][0]["valor_completo"],
            "R$ 5.000,00",
        )
        self.assertEqual(
            response.context["faturamento"]["indicadores"][1]["valor_completo"],
            "R$ 10.000,00",
        )
        self.assertEqual(response.context["vendedores_selecionados"], [f"{self.empresa.pk}:{vendedor.codigo}"])
        self.assertEqual(
            response.context["tipos_faturamento_selecionados"],
            ["produtos", "servicos"],
        )

        response = self.client.get(
            reverse(
                "dashboards:exportar_faturamento_produtos",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
                "vendedor": f"{self.empresa.pk}:{vendedor.codigo}",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            [
                "Data de Emissao",
                "Cliente",
                "Numero da NF",
                "Total de Mercadoria",
                "Frete",
                "Total da Nota Fiscal",
            ],
        )
        self.assertEqual(worksheet["B2"].value, "Cliente Excel")
        self.assertEqual(worksheet["C2"].value, "NF-9001")
        self.assertEqual(worksheet["D2"].value, 2700)
        self.assertEqual(worksheet["E2"].value, 300)
        self.assertEqual(worksheet["F2"].value, 3000)

        response = self.client.get(
            reverse(
                "dashboards:exportar_faturamento_servicos",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
                "vendedor": f"{self.empresa.pk}:{vendedor.codigo}",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            [
                "Data de Emissao",
                "Cliente",
                "Numero da NFS-e ou Recibo",
                "Total da Nota Fiscal",
            ],
        )
        self.assertEqual(worksheet["B2"].value, "Cliente Excel")
        self.assertEqual(worksheet["C2"].value, "8859")
        self.assertEqual(worksheet["D2"].value, 1200)
        self.assertEqual(worksheet["B3"].value, "Cliente Excel")
        self.assertEqual(worksheet["C3"].value, "456")
        self.assertEqual(worksheet["D3"].value, 800)
        self.assertIsNone(worksheet["A4"].value)

    def test_faturamento_linha_faturado_usa_total_mensal(self):
        ano_atual = date.today().year
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=310,
            numero_pedido="PV-310",
            data_inclusao=date(ano_atual, 1, 5),
            data_faturamento=date(ano_atual, 1, 8),
            faturado=True,
            valor_total_pedido=1000,
        )
        ordem_janeiro = OrdemServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_os=410,
            numero_os="OS-410",
            data_inclusao=date(ano_atual, 1, 7),
            data_faturamento=date(ano_atual, 1, 9),
            faturada=True,
            valor_total=300,
        )
        NfseOmie.objects.create(
            empresa=self.empresa,
            codigo_nf=9010,
            numero_nfse="9010",
            status_nfse="F",
            codigo_os=ordem_janeiro.codigo_os,
            numero_os=ordem_janeiro.numero_os,
            ordem_servico=ordem_janeiro,
            data_emissao=ordem_janeiro.data_faturamento,
            valor_nfse=300,
        )
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=320,
            numero_pedido="PV-320",
            data_inclusao=date(ano_atual, 2, 5),
            data_faturamento=date(ano_atual, 2, 8),
            faturado=True,
            valor_total_pedido=2000,
        )
        ordem_fevereiro = OrdemServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_os=420,
            numero_os="OS-420",
            data_inclusao=date(ano_atual, 2, 7),
            data_faturamento=date(ano_atual, 2, 9),
            faturada=True,
            valor_total=400,
        )
        NfseOmie.objects.create(
            empresa=self.empresa,
            codigo_nf=9020,
            numero_nfse="9020",
            status_nfse="F",
            codigo_os=ordem_fevereiro.codigo_os,
            numero_os=ordem_fevereiro.numero_os,
            ordem_servico=ordem_fevereiro,
            data_emissao=ordem_fevereiro.data_faturamento,
            valor_nfse=400,
        )

        contexto = faturamento_comercial(
            self.empresa,
            f"tri-{ano_atual}-1",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(contexto["acumulado"], [1300.0, 2400.0, 0.0])

    def test_faturamento_produtos_considera_impostos_do_json_omie(self):
        ano_atual = date.today().year
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=330,
            numero_pedido="PV-330",
            data_inclusao=date(ano_atual, 1, 5),
            data_faturamento=date(ano_atual, 1, 8),
            faturado=True,
            valor_mercadorias=1000,
            valor_frete=100,
            valor_seguro=50,
            frete={"outras_despesas": 25},
            valor_total_pedido=1175,
            dados_originais={
                "total_pedido": {
                    "valor_icms": 120,
                    "valor_pis": 15,
                    "valor_cofins": 35,
                }
            },
        )

        contexto = faturamento_comercial(
            self.empresa,
            f"mes-{ano_atual}-01",
            empresas_ids=[self.empresa.pk],
            tipos_selecionados=["impostos"],
        )

        self.assertEqual(contexto["produtos_mercadorias"], [0])
        self.assertEqual(contexto["produtos_frete_despesas"], [0])
        self.assertEqual(contexto["produtos_impostos"], [170.0])
        self.assertEqual(contexto["servicos_impostos"], [0])
        self.assertEqual(contexto["acumulado"], [170.0])
        self.assertEqual(contexto["indicadores"][0]["valor_completo"], "R$ 170,00")

    def test_faturamento_impostos_considera_servicos_sem_retencao_e_iss(self):
        ano_atual = date.today().year
        ordem = OrdemServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_os=430,
            numero_os="OS-430",
            data_inclusao=date(ano_atual, 1, 5),
            data_faturamento=date(ano_atual, 1, 9),
            faturada=True,
            valor_total=1000,
        )
        NfseOmie.objects.create(
            empresa=self.empresa,
            codigo_nf=9030,
            numero_nfse="9030",
            status_nfse="F",
            codigo_os=ordem.codigo_os,
            numero_os=ordem.numero_os,
            ordem_servico=ordem,
            data_emissao=ordem.data_faturamento,
            valor_nfse=1000,
        )
        OrdemServicoItemOmie.objects.create(
            empresa=self.empresa,
            ordem_servico=ordem,
            codigo_item=431,
            descricao="Servico tributado",
            quantidade=1,
            valor_unitario=1000,
            valor_iss=50,
            impostos={
                "nValorCOFINS": 30,
                "nValorCSLL": 10,
                "nValorINSS": 40,
                "nValorIRRF": 15,
                "nValorISS": 50,
                "nValorPIS": 7,
                "cRetemCOFINS": "N",
                "cRetemCSLL": "S",
                "cRetemINSS": "N",
                "cRetemIRRF": "S",
                "cRetemPIS": "N",
            },
        )

        contexto = faturamento_comercial(
            self.empresa,
            f"mes-{ano_atual}-01",
            empresas_ids=[self.empresa.pk],
            tipos_selecionados=["impostos"],
        )

        self.assertEqual(contexto["produtos_impostos"], [0])
        self.assertEqual(contexto["servicos_impostos"], [127.0])
        self.assertEqual(contexto["acumulado"], [127.0])
        self.assertEqual(contexto["indicadores"][0]["valor_completo"], "R$ 127,00")

    def test_desempenho_vendedores_exibe_indicadores_graficos_e_carteira(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        ana = VendedorOmie.objects.create(
            empresa=self.empresa,
            codigo=201,
            nome="Ana",
        )
        carlos = VendedorOmie.objects.create(
            empresa=self.empresa,
            codigo=202,
            nome="Carlos",
        )
        fernanda = VendedorOmie.objects.create(
            empresa=self.empresa,
            codigo=203,
            nome="Fernanda",
        )
        for vendedor, meta in ((ana, 10000), (carlos, 9000), (fernanda, 8000)):
            MetaVendedorComercial.objects.create(
                empresa=self.empresa,
                vendedor=vendedor,
                ano=ano_atual,
                mes=1,
                valor_mensal=meta,
            )
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=2101,
            numero_pedido="PV-2101",
            codigo_vendedor=ana.codigo,
            data_inclusao=date(ano_atual, 1, 5),
            data_faturamento=date(ano_atual, 1, 10),
            faturado=True,
            valor_total_pedido=12000,
        )
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=2102,
            numero_pedido="PV-2102",
            codigo_vendedor=carlos.codigo,
            data_inclusao=date(ano_atual, 1, 6),
            data_faturamento=date(ano_atual, 1, 11),
            faturado=True,
            valor_total_pedido=7000,
        )
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=2103,
            numero_pedido="PV-2103",
            codigo_vendedor=fernanda.codigo,
            data_inclusao=date(ano_atual, 1, 7),
            data_faturamento=date(ano_atual, 1, 12),
            faturado=True,
            valor_total_pedido=3000,
        )
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=2201,
            numero_pedido="PV-2201",
            codigo_vendedor=ana.codigo,
            data_inclusao=date(ano_atual, 1, 15),
            faturado=False,
            valor_total_pedido=5000,
        )
        ordem_ana = OrdemServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_os=2301,
            numero_os="OS-2301",
            codigo_vendedor=ana.codigo,
            data_inclusao=date(ano_atual, 1, 16),
            data_faturamento=date(ano_atual, 1, 20),
            faturada=True,
            valor_total=4000,
        )
        NfseOmie.objects.create(
            empresa=self.empresa,
            codigo_nf=9231,
            numero_nfse="9231",
            status_nfse="F",
            codigo_os=ordem_ana.codigo_os,
            numero_os=ordem_ana.numero_os,
            ordem_servico=ordem_ana,
            data_emissao=ordem_ana.data_faturamento,
            valor_nfse=4000,
        )
        OrdemServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_os=2302,
            numero_os="OS-2302",
            codigo_vendedor=carlos.codigo,
            data_inclusao=date(ano_atual, 1, 18),
            faturada=False,
            valor_total=2500,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "desempenho-de-vendedores",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
                "tipo_faturamento": ["produtos", "servicos"],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Faturamento total da equipe")
        self.assertContains(response, "% da meta atingida")
        self.assertContains(response, "Melhor performance")
        self.assertContains(response, "Ticket medio da equipe")
        self.assertContains(response, "Meta vs. realizado")
        self.assertContains(response, "Tendencia mensal - top 3 vendedores")
        self.assertContains(response, "Pedidos em aberto por vendedor")
        self.assertContains(response, "Produtos")
        self.assertContains(response, "Servicos")
        self.assertContains(response, "Ana")
        self.assertContains(response, "Carlos")
        self.assertContains(response, "Fernanda")
        self.assertContains(response, "Acima da meta")
        self.assertContains(response, "data-seller-ranking-chart")
        self.assertContains(response, "data-seller-trend-chart")
        self.assertEqual(
            response.context["tipos_faturamento_selecionados"],
            ["produtos", "servicos"],
        )

    def test_analise_clientes_exibe_indicadores_e_graficos(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        vip = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=3101,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente VIP",
        )
        ativo = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=3102,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente Ativo",
        )
        risco = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=3103,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente Risco",
        )
        CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=3104,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente Inativo",
            inativo=True,
        )
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=3201,
            numero_pedido="PV-3201",
            codigo_cliente=vip.codigo_cliente_omie,
            cliente=vip,
            data_faturamento=date(ano_atual, 1, 10),
            faturado=True,
            valor_total_pedido=5000,
        )
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=3202,
            numero_pedido="PV-3202",
            codigo_cliente=vip.codigo_cliente_omie,
            cliente=vip,
            data_faturamento=date(ano_atual, 2, 10),
            faturado=True,
            valor_total_pedido=3000,
        )
        ordem_ativo = OrdemServicoOmie.objects.create(
            empresa=self.empresa,
            codigo_os=3301,
            numero_os="OS-3301",
            codigo_cliente=ativo.codigo_cliente_omie,
            cliente=ativo,
            data_faturamento=date(ano_atual, 3, 10),
            faturada=True,
            valor_total=2000,
        )
        NfseOmie.objects.create(
            empresa=self.empresa,
            codigo_nf=9331,
            numero_nfse="9331",
            status_nfse="F",
            codigo_os=ordem_ativo.codigo_os,
            numero_os=ordem_ativo.numero_os,
            ordem_servico=ordem_ativo,
            data_emissao=ordem_ativo.data_faturamento,
            valor_nfse=2000,
        )
        PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=3203,
            numero_pedido="PV-3203",
            codigo_cliente=risco.codigo_cliente_omie,
            cliente=risco,
            data_faturamento=date(ano_atual - 1, 9, 10),
            faturado=True,
            valor_total_pedido=1500,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "analise-de-clientes",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"ano-{ano_atual}",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Clientes ativos")
        self.assertContains(response, "Novos no periodo")
        self.assertContains(response, "Churn no periodo")
        self.assertContains(response, "Ticket medio")
        self.assertContains(response, "Tipos de clientes")
        self.assertContains(response, "Top 10 clientes vs. restante")
        self.assertContains(response, "Ticket medio por segmento")
        self.assertContains(response, "Cliente VIP")
        self.assertContains(response, "data-client-segment-chart")
        self.assertContains(response, "data-client-top-chart")
        self.assertContains(response, "data-client-ticket-chart")

    def test_margem_rentabilidade_exibe_indicadores_grafico_e_rankings(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        produto_lucrativo = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=4101,
            codigo="LUC-01",
            descricao="Produto Lucrativo",
            info={"valor_custo": "40"},
        )
        produto_critico = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=4102,
            codigo="CRT-01",
            descricao="Produto Critico",
            info={"valor_custo": "130"},
        )
        PosicaoEstoqueOmie.objects.create(
            empresa=self.empresa,
            produto=produto_lucrativo,
            codigo_produto=produto_lucrativo.codigo_produto,
            codigo_local_estoque=0,
            codigo=produto_lucrativo.codigo,
            descricao=produto_lucrativo.descricao,
            cmc=40,
        )
        PosicaoEstoqueOmie.objects.create(
            empresa=self.empresa,
            produto=produto_critico,
            codigo_produto=produto_critico.codigo_produto,
            codigo_local_estoque=0,
            codigo=produto_critico.codigo,
            descricao=produto_critico.descricao,
            cmc=130,
        )
        pedido = PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=4201,
            numero_pedido="PV-4201",
            data_faturamento=date(ano_atual, 1, 15),
            faturado=True,
            valor_total_pedido=2400,
            valor_descontos=100,
        )
        PedidoItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido,
            codigo_item=4301,
            produto=produto_lucrativo,
            codigo_produto=produto_lucrativo.codigo_produto,
            codigo_produto_texto=produto_lucrativo.codigo,
            descricao=produto_lucrativo.descricao,
            quantidade=10,
            valor_unitario=150,
            valor_total=1500,
            valor_desconto=50,
        )
        PedidoItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido,
            codigo_item=4302,
            produto=produto_critico,
            codigo_produto=produto_critico.codigo_produto,
            codigo_produto_texto=produto_critico.codigo,
            descricao=produto_critico.descricao,
            quantidade=10,
            valor_unitario=90,
            valor_total=900,
            valor_desconto=50,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "margem-e-rentabilidade",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Margem bruta media")
        self.assertContains(response, "Receita total")
        self.assertContains(response, "Produtos c/ mg negativa")
        self.assertContains(response, "Desconto medio")
        self.assertContains(response, "Receita x margem")
        self.assertContains(response, "Top 5 mais rentaveis")
        self.assertContains(response, "Bottom 5 - Atencao urgente")
        self.assertContains(response, "Produto Lucrativo")
        self.assertContains(response, "Produto Critico")
        self.assertContains(response, "data-margin-bubble-chart")

    def test_margem_rentabilidade_usa_cmc_e_valor_unitario_dos_pedidos(self):
        ano_atual = date.today().year
        produto = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=5101,
            codigo="CMC-01",
            descricao="Produto CMC",
            info={"valor_custo": "1"},
        )
        produto_sem_cmc = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=5102,
            codigo="SEM-CMC",
            descricao="Produto Sem CMC",
        )
        produto_cmc_zero = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=5103,
            codigo="CMC-ZERO",
            descricao="Produto CMC Zero",
        )
        PosicaoEstoqueOmie.objects.create(
            empresa=self.empresa,
            produto=produto,
            codigo_produto=produto.codigo_produto,
            codigo_local_estoque=123,
            codigo=produto.codigo,
            descricao=produto.descricao,
            cmc=30,
        )
        PosicaoEstoqueOmie.objects.create(
            empresa=self.empresa,
            produto=produto_cmc_zero,
            codigo_produto=produto_cmc_zero.codigo_produto,
            codigo_local_estoque=123,
            codigo=produto_cmc_zero.codigo,
            descricao=produto_cmc_zero.descricao,
            cmc=0,
        )
        pedido = PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=5201,
            numero_pedido="PV-5201",
            data_faturamento=date(ano_atual, 1, 15),
            faturado=True,
            valor_total_pedido=9999,
        )
        PedidoItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido,
            codigo_item=5301,
            produto=produto,
            codigo_produto=produto.codigo_produto,
            codigo_produto_texto=produto.codigo,
            codigo_local_estoque=123,
            descricao=produto.descricao,
            quantidade=10,
            valor_unitario=50,
            valor_total=9999,
        )
        PedidoItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido,
            codigo_item=5302,
            produto=produto_sem_cmc,
            codigo_produto=produto_sem_cmc.codigo_produto,
            codigo_produto_texto=produto_sem_cmc.codigo,
            codigo_local_estoque=123,
            descricao=produto_sem_cmc.descricao,
            quantidade=10,
            valor_unitario=100,
            valor_total=1000,
        )
        PedidoItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido,
            codigo_item=5303,
            produto=produto_cmc_zero,
            codigo_produto=produto_cmc_zero.codigo_produto,
            codigo_produto_texto=produto_cmc_zero.codigo,
            codigo_local_estoque=123,
            descricao=produto_cmc_zero.descricao,
            quantidade=10,
            valor_unitario=100,
            valor_total=1000,
        )

        contexto = margem_rentabilidade_comercial(
            self.empresa,
            f"mes-{ano_atual}-01",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(contexto["indicadores"][0]["valor"], "40.0%")
        self.assertEqual(contexto["indicadores"][1]["valor"], "R$ 500,00")
        self.assertEqual(len(contexto["top_rentaveis"]), 1)
        self.assertEqual(contexto["top_rentaveis"][0]["receita"], Decimal("500.0000"))
        self.assertEqual(contexto["top_rentaveis"][0]["margem_fmt"], "40.0%")

    def test_dashboard_permite_multisselecao_de_filtros(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        outra_empresa = Empresa.objects.create(
            nome="Filial Oeste Ltda",
            nome_fantasia="Filial Oeste",
            cnpj="22.222.222/0001-22",
            grupo="Grupo Oeste",
        )
        EmpresaUsuario.objects.create(empresa=outra_empresa, usuario=self.usuario)
        projeto_principal = ProjetoOmie.objects.create(
            empresa=self.empresa,
            codigo=101,
            nome="Projeto principal",
        )
        projeto_filial = ProjetoOmie.objects.create(
            empresa=outra_empresa,
            codigo=202,
            nome="Projeto filial",
        )
        departamento_principal = DepartamentoOmie.objects.create(
            empresa=self.empresa,
            codigo="DEP-1",
            descricao="Comercial matriz",
        )
        departamento_filial = DepartamentoOmie.objects.create(
            empresa=outra_empresa,
            codigo="DEP-2",
            descricao="Comercial filial",
        )
        self.client.force_login(self.usuario)
        self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug}),
            {
                "_filtrar_empresas": "1",
                "empresa": [str(self.empresa.pk), str(outra_empresa.pk)],
            },
        )
        url = reverse(
            "dashboards:dashboard",
            kwargs={
                "empresa_slug": self.empresa.slug,
                "area_slug": "comercial",
                "dashboard_slug": "faturamento",
            },
        )

        response = self.client.get(
            url,
            {
                "_filtrar": "1",
                "projeto": [
                    f"{self.empresa.pk}:{projeto_principal.codigo}",
                    f"{outra_empresa.pk}:{projeto_filial.codigo}",
                ],
                "departamento": [
                    f"{self.empresa.pk}:{departamento_principal.codigo}",
                    f"{outra_empresa.pk}:{departamento_filial.codigo}",
                ],
                "periodo": f"tri-{ano_atual}-2",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["empresas_selecionadas"]), 2)
        self.assertEqual(len(response.context["projetos_selecionados"]), 2)
        self.assertEqual(len(response.context["departamentos_selecionados"]), 2)
        self.assertContains(response, "Projeto principal")
        self.assertContains(response, "Projeto filial")
        self.assertContains(response, "Comercial matriz")
        self.assertContains(response, "Comercial filial")
        self.assertNotContains(response, "Todas as empresas")

    def test_select_de_empresas_exibe_apenas_empresas_do_mesmo_grupo(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        mesma_rede = Empresa.objects.create(
            nome="Filial Mesmo Grupo Ltda",
            nome_fantasia="Filial Mesmo Grupo",
            cnpj="22.222.222/0001-22",
            grupo="grupo oeste",
        )
        outro_grupo = Empresa.objects.create(
            nome="Cliente Outro Grupo Ltda",
            nome_fantasia="Cliente Outro Grupo",
            cnpj="33.333.333/0001-33",
            grupo="Grupo Leste",
        )
        EmpresaUsuario.objects.create(empresa=mesma_rede, usuario=self.usuario)
        EmpresaUsuario.objects.create(empresa=outro_grupo, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug}),
            {
                "_filtrar_empresas": "1",
                "empresa": [
                    str(self.empresa.pk),
                    str(mesma_rede.pk),
                    str(outro_grupo.pk),
                ],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context["empresas_selecionadas"],
            [str(self.empresa.pk), str(mesma_rede.pk)],
        )
        self.assertContains(response, "Filial Mesmo Grupo")
        self.assertNotContains(response, "Cliente Outro Grupo")

    def test_empresa_sem_grupo_exibe_apenas_ela_no_select(self):
        self.empresa.grupo = ""
        self.empresa.save(update_fields=["grupo"])
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        outra_empresa = Empresa.objects.create(
            nome="Outra Empresa Sem Grupo Ltda",
            nome_fantasia="Outra Empresa Sem Grupo",
            cnpj="44.444.444/0001-44",
        )
        EmpresaUsuario.objects.create(empresa=outra_empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse("dashboards:home", kwargs={"empresa_slug": self.empresa.slug}),
            {
                "_filtrar_empresas": "1",
                "empresa": [str(self.empresa.pk), str(outra_empresa.pk)],
            },
        )

        self.assertEqual(response.context["empresas_selecionadas"], [str(self.empresa.pk)])
        self.assertContains(response, self.empresa.nome_fantasia)
        self.assertNotContains(response, outra_empresa.nome_fantasia)

    def test_limpar_filtros_marca_todas_as_opcoes_do_dashboard(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        projeto_a = ProjetoOmie.objects.create(
            empresa=self.empresa,
            codigo=101,
            nome="Projeto A",
        )
        projeto_b = ProjetoOmie.objects.create(
            empresa=self.empresa,
            codigo=202,
            nome="Projeto B",
        )
        departamento_a = DepartamentoOmie.objects.create(
            empresa=self.empresa,
            codigo="DEP-A",
            descricao="Departamento A",
        )
        departamento_b = DepartamentoOmie.objects.create(
            empresa=self.empresa,
            codigo="DEP-B",
            descricao="Departamento B",
        )
        produto_sem_projeto = ProdutoOmie.objects.create(
            empresa=self.empresa,
            codigo_produto=909,
            codigo="SEM-PROJ",
            descricao="Produto sem projeto",
        )
        pedido_sem_projeto = PedidoOmie.objects.create(
            empresa=self.empresa,
            codigo_pedido=909,
            numero_pedido="PV-SEM-PROJ",
            data_inclusao=date.today(),
            data_faturamento=date.today(),
            faturado=True,
            valor_total_pedido=1000,
        )
        PedidoItemOmie.objects.create(
            empresa=self.empresa,
            pedido=pedido_sem_projeto,
            codigo_item=909,
            produto=produto_sem_projeto,
            codigo_produto=produto_sem_projeto.codigo_produto,
            descricao=produto_sem_projeto.descricao,
            quantidade=1,
            valor_unitario=1000,
            valor_total=1000,
        )
        self.client.force_login(self.usuario)
        url = reverse(
            "dashboards:dashboard",
            kwargs={
                "empresa_slug": self.empresa.slug,
                "area_slug": "comercial",
                "dashboard_slug": "faturamento",
            },
        )

        response = self.client.get(
            url,
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-02",
                "regime_financeiro": "competencia",
                "projeto": [f"{self.empresa.pk}:{projeto_a.codigo}"],
                "departamento": [f"{self.empresa.pk}:{departamento_a.codigo}"],
            },
        )
        self.assertEqual(response.context["projetos_selecionados"], [f"{self.empresa.pk}:{projeto_a.codigo}"])
        self.assertEqual(response.context["departamentos_selecionados"], [f"{self.empresa.pk}:{departamento_a.codigo}"])
        self.assertNotContains(response, "Produto sem projeto")

        response = self.client.get(
            url,
            {
                "_filtrar": "1",
                "limpar_filtros": "1",
                "periodo": f"mes-{ano_atual}-02",
                "regime_financeiro": "competencia",
            },
        )

        self.assertContains(response, "Limpar filtros")
        self.assertContains(response, "Produto sem projeto")
        self.assertEqual(response.context["periodo_selecionado"], f"ano-{ano_atual}")
        self.assertEqual(response.context["periodo_rotulo"], str(ano_atual))
        self.assertEqual(response.context["regime_financeiro"], "caixa")
        self.assertContains(response, "Todos os projetos")
        self.assertContains(response, "Todos os departamentos")
        self.assertEqual(
            set(response.context["projetos_selecionados"]),
            {
                f"{self.empresa.pk}:{projeto_a.codigo}",
                f"{self.empresa.pk}:{projeto_b.codigo}",
            },
        )
        self.assertEqual(
            set(response.context["departamentos_selecionados"]),
            {
                f"{self.empresa.pk}:{departamento_a.codigo}",
                f"{self.empresa.pk}:{departamento_b.codigo}",
            },
        )

    def test_periodo_pode_ser_compartilhado_no_modulo(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)
        url_vendas = reverse(
            "dashboards:dashboard",
            kwargs={
                "empresa_slug": self.empresa.slug,
                "area_slug": "comercial",
                "dashboard_slug": "faturamento",
            },
        )
        url_clientes = reverse(
            "dashboards:dashboard",
            kwargs={
                "empresa_slug": self.empresa.slug,
                "area_slug": "comercial",
                "dashboard_slug": "analise-de-clientes",
            },
        )

        response = self.client.get(
            url_vendas,
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-04",
                "compartilhar_periodo": "1",
            },
        )
        self.assertTrue(response.context["periodo_foi_compartilhado"])

        response = self.client.get(url_clientes)
        self.assertEqual(
            response.context["periodo_selecionado"],
            f"mes-{ano_atual}-04",
        )
        self.assertTrue(response.context["periodo_compartilhado"])
        self.assertContains(response, f"Abril de {ano_atual}")

    def test_filtro_de_data_exibe_anos_trimestres_e_meses(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "faturamento",
                },
            )
        )

        self.assertContains(response, "1º trimestre")
        self.assertContains(response, "4º trimestre")
        self.assertContains(response, "Janeiro")
        self.assertContains(response, "Dezembro")
        self.assertContains(response, "Período específico")
        self.assertContains(response, "Usar no módulo")

    def test_periodo_especifico_e_compartilhado_no_modulo(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)
        url_vendas = reverse(
            "dashboards:dashboard",
            kwargs={
                "empresa_slug": self.empresa.slug,
                "area_slug": "comercial",
                "dashboard_slug": "faturamento",
            },
        )
        url_produtos = reverse(
            "dashboards:dashboard",
            kwargs={
                "empresa_slug": self.empresa.slug,
                "area_slug": "comercial",
                "dashboard_slug": "margem-e-rentabilidade",
            },
        )

        response = self.client.get(
            url_vendas,
            {
                "_filtrar": "1",
                "periodo": "personalizado",
                "data_inicio": "2026-01-15",
                "data_fim": "2026-02-20",
                "compartilhar_periodo": "1",
            },
        )

        self.assertEqual(response.context["periodo_selecionado"], "personalizado")
        self.assertEqual(response.context["data_inicio"], "2026-01-15")
        self.assertEqual(response.context["data_fim"], "2026-02-20")
        self.assertContains(response, "15/01/2026 a 20/02/2026")

        response = self.client.get(url_produtos)
        self.assertEqual(response.context["periodo_selecionado"], "personalizado")
        self.assertEqual(response.context["data_inicio"], "2026-01-15")
        self.assertEqual(response.context["data_fim"], "2026-02-20")

    def test_periodo_especifico_invalido_volta_ao_ano_atual(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "faturamento",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": "personalizado",
                "data_inicio": "2026-03-10",
                "data_fim": "2026-03-01",
            },
        )

        self.assertEqual(
            response.context["periodo_selecionado"],
            f"ano-{date.today().year}",
        )

    def test_dashboard_inexistente_retorna_404(self):
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "comercial",
                    "dashboard_slug": "inexistente",
                },
            )
        )

        self.assertEqual(response.status_code, 404)

    def test_dre_gerencial_exibe_indicadores_tabela_e_graficos(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        receita = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receita Bruta",
            ordem=1,
        )
        custos = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Deducoes e Gastos Variaveis",
            sinal=ContaDRE.Sinal.SUBTRACAO,
            ordem=2,
        )
        custos_variaveis = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Custos Variaveis",
            conta_pai=custos,
            sinal=ContaDRE.Sinal.SUBTRACAO,
            ordem=1,
        )
        categoria = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.02.02",
            descricao="Materia prima",
            conta_dre=custos_variaveis,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=1001,
            data_lancamento=date(ano_atual, 1, 10),
            valor_lancamento=1200000,
            categoria_principal=categoria,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                    "dashboard_slug": "dre-gerencial",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
                "regime_financeiro": "caixa",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Receita bruta")
        self.assertContains(response, "Deducoes e gastos variaveis")
        self.assertContains(response, "DRE Gerencial")
        self.assertContains(response, "Custos Variaveis")
        self.assertContains(response, "R$ -1.20 Mi")
        self.assertContains(response, 'title="R$ -1.200.000,00"')
        self.assertContains(response, "data-dre-expand-all")
        self.assertContains(response, "data-dre-values-chart")
        self.assertContains(response, "data-dre-ah-chart")
        self.assertContains(response, 'option value="caixa" selected')

    def test_visao_geral_financeira_exibe_indicadores_graficos_e_rankings(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        cliente = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=2001,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente Forte",
            razao_social="Cliente Forte Ltda",
        )
        fornecedor = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=3001,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="Fornecedor Chave",
            razao_social="Fornecedor Chave Ltda",
        )
        categoria_receita = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.01",
            descricao="Servicos Prestados",
        )
        categoria_despesa = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.01.01",
            descricao="Alimentacao",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=4001,
            cliente=cliente,
            categoria_principal=categoria_receita,
            data_previsao=date(ano_atual, 1, 10),
            valor_documento=5000,
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=5001,
            fornecedor=fornecedor,
            categoria_principal=categoria_despesa,
            data_previsao=date(ano_atual, 1, 12),
            valor_documento=1800,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                    "dashboard_slug": "visao-geral",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
                "regime_financeiro": "competencia",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Recebimentos")
        self.assertContains(response, "Pagamentos")
        self.assertContains(response, "Resultado")
        self.assertContains(response, 'title="R$ 5.000,00"')
        self.assertContains(response, "Margem mensal")
        self.assertContains(response, "Cliente Forte")
        self.assertContains(response, "Fornecedor Chave")
        self.assertContains(response, "data-overview-flow-chart")
        self.assertContains(response, "data-overview-margin-chart")
        self.assertContains(response, 'option value="competencia" selected')

    def test_visao_geral_filtra_por_categoria_financeira(self):
        categoria_pai = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01",
            descricao="Receitas",
            totalizadora=True,
        )
        categoria_selecionada = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.01",
            categoria_superior=categoria_pai.codigo,
            descricao="Receita selecionada",
        )
        categoria_fora = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.02",
            categoria_superior=categoria_pai.codigo,
            descricao="Receita fora",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=88001,
            codigo_categoria=categoria_selecionada.codigo,
            categoria_principal=categoria_selecionada,
            data_registro=date.today(),
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_documento=100,
            valor_a_receber=100,
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=88002,
            codigo_categoria=categoria_fora.codigo,
            categoria_principal=categoria_fora,
            data_registro=date.today(),
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_documento=900,
            valor_a_receber=900,
        )

        contexto = visao_geral_financeira(
            self.empresa,
            f"mes-{date.today().year}-{date.today().month:02d}",
            empresas_ids=[self.empresa.pk],
            regime_financeiro="competencia",
            categorias_selecionadas=[
                f"{self.empresa.pk}:{categoria_selecionada.codigo}"
            ],
        )

        self.assertEqual(contexto["indicadores"][0]["valor_completo"], "R$ 100,00")
        self.assertEqual(contexto["indicadores"][2]["valor_completo"], "R$ 100,00")

    def test_visao_geral_caixa_usa_lancamentos_de_conta_corrente(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        cliente = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=6001,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente Caixa",
        )
        fornecedor = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=7001,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="Fornecedor Caixa",
        )
        categoria_receita = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.02",
            descricao="Mensalidade",
        )
        categoria_despesa = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.01.02",
            descricao="Insumos",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=8001,
            cliente=cliente,
            categoria_principal=categoria_receita,
            data_registro=date(ano_atual, 2, 10),
            valor_documento=9900,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=9001,
            cliente_fornecedor=cliente,
            categoria_principal=categoria_receita,
            data_lancamento=date(ano_atual, 1, 10),
            natureza="R",
            valor_lancamento=1200,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=9002,
            cliente_fornecedor=fornecedor,
            categoria_principal=categoria_despesa,
            data_lancamento=date(ano_atual, 1, 11),
            natureza="P",
            valor_lancamento=400,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                    "dashboard_slug": "visao-geral",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
                "regime_financeiro": "caixa",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Cliente Caixa")
        self.assertContains(response, "Fornecedor Caixa")
        self.assertNotContains(response, "R$ 9.900,00")

    def test_fluxo_de_caixa_exibe_indicadores_grafico_e_composicoes_realizadas(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        cliente = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=10001,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente Critico",
        )
        fornecedor = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=10002,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="Fornecedor Critico",
        )
        fornecedor_reagendado = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=10003,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="Fornecedor Reagendado",
        )
        categoria_receita_pai = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.03",
            descricao="Receitas recorrentes",
        )
        categoria_receita = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.03.01",
            categoria_superior="1.03",
            descricao="Assinaturas",
        )
        categoria_despesa_pai = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.03",
            descricao="Despesas operacionais",
        )
        categoria_despesa = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.03.01",
            categoria_superior="2.03",
            descricao="Operacional",
        )
        ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=101,
            descricao="Conta principal",
            saldo_inicial=15000,
            saldo_atual=15000,
        )
        projeto_horizontal = ProjetoOmie.objects.create(
            empresa=self.empresa,
            codigo=90901,
            nome="Projeto horizontal",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=11001,
            cliente=cliente,
            categoria_principal=categoria_receita,
            data_previsao=date(ano_atual, 1, 10),
            data_vencimento=date(ano_atual, 1, 10),
            valor_documento=7000,
            valor_a_receber=7000,
            status_titulo="ATRASADO",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=12001,
            fornecedor=fornecedor,
            categoria_principal=categoria_despesa,
            data_entrada=date(ano_atual, 1, 1),
            data_previsao=date(ano_atual, 1, 15),
            data_vencimento=date(ano_atual, 1, 15),
            valor_documento=2500,
            valor_a_pagar=2500,
            status_titulo="ATRASADO",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=12002,
            fornecedor=fornecedor_reagendado,
            categoria_principal=categoria_despesa,
            data_entrada=date.today() - timedelta(days=20),
            data_previsao=date.today() + timedelta(days=10),
            data_vencimento=date.today() - timedelta(days=5),
            valor_documento=800,
            valor_a_pagar=800,
            status_titulo="ATRASADO",
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=13001,
            cliente_fornecedor=cliente,
            categoria_principal=categoria_receita,
            data_lancamento=date(ano_atual, 1, 20),
            natureza="R",
            valor_lancamento=900,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=13002,
            cliente_fornecedor=fornecedor,
            categoria_principal=categoria_despesa,
            data_lancamento=date(ano_atual, 1, 21),
            natureza="P",
            valor_lancamento=300,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=14001,
            cliente_fornecedor=cliente,
            codigo_conta_corrente=101,
            codigo_categoria=categoria_receita.codigo,
            categoria_principal=categoria_receita,
            grupo="CONTA_A_RECEBER",
            natureza="R",
            status="ABERTO",
            data_previsao=date(ano_atual, 1, 10),
            data_vencimento=date(ano_atual, 1, 10),
            valor_titulo=7000,
            valor_liquido=7000,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=14002,
            cliente_fornecedor=cliente,
            codigo_conta_corrente=101,
            codigo_categoria=categoria_receita.codigo,
            categoria_principal=categoria_receita,
            grupo="CONTA_A_RECEBER",
            natureza="R",
            status="PAGO",
            data_previsao=date(ano_atual, 1, 20),
            data_pagamento=date(ano_atual, 1, 20),
            data_vencimento=date(ano_atual, 1, 20),
            valor_titulo=900,
            valor_liquido=900,
            valor_pago=900,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=14003,
            cliente_fornecedor=fornecedor,
            codigo_conta_corrente=101,
            codigo_categoria=categoria_despesa.codigo,
            categoria_principal=categoria_despesa,
            grupo="CONTA_A_PAGAR",
            natureza="P",
            status="ABERTO",
            data_previsao=date(ano_atual, 1, 15),
            data_vencimento=date(ano_atual, 1, 15),
            valor_titulo=2500,
            valor_liquido=2500,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=14005,
            cliente_fornecedor=fornecedor_reagendado,
            codigo_conta_corrente=101,
            codigo_categoria=categoria_despesa.codigo,
            categoria_principal=categoria_despesa,
            grupo="CONTA_A_PAGAR",
            natureza="P",
            status="ABERTO",
            liquidado=False,
            data_previsao=date.today() + timedelta(days=10),
            data_vencimento=date.today() - timedelta(days=5),
            valor_titulo=800,
            valor_aberto=800,
            valor_liquido=800,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=14004,
            cliente_fornecedor=fornecedor,
            codigo_conta_corrente=101,
            codigo_categoria=categoria_despesa.codigo,
            categoria_principal=categoria_despesa,
            grupo="CONTA_A_PAGAR",
            natureza="P",
            status="PAGO",
            data_previsao=date(ano_atual, 1, 21),
            data_vencimento=date(ano_atual, 1, 21),
            valor_titulo=300,
            valor_liquido=300,
            valor_pago=300,
        )
        self.client.force_login(self.usuario)
        session = self.client.session
        session[f"filtros_dashboard:{self.empresa.pk}:financeiro:fluxo-de-caixa"] = {
            "projetos": [f"{self.empresa.pk}:{projeto_horizontal.codigo}"],
        }
        session.save()

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                    "dashboard_slug": "fluxo-de-caixa",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"mes-{ano_atual}-01",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Saldo atual")
        self.assertContains(response, "Entradas previstas")
        self.assertContains(response, "Saidas previstas")
        self.assertContains(response, 'title="R$ 15.000,00"')
        self.assertContains(response, "Saldo projetado")
        self.assertContains(response, "Prazo medio de recebimento")
        self.assertContains(response, "Recebimentos pendentes - criticos")
        self.assertContains(response, "Cliente Critico")
        self.assertContains(response, "Fornecedor Critico")
        self.assertContains(response, "Composicao das entradas realizadas")
        self.assertContains(response, "Composicao das saidas realizadas")
        self.assertNotContains(response, "Fornecedor Reagendado")
        self.assertContains(response, "data-cashflow-chart")
        self.assertContains(response, "data-cashflow-in-pie")
        self.assertContains(response, "data-cashflow-out-pie")
        self.assertContains(response, "Quer detalhar as contas?")
        self.assertContains(response, "data-cashflow-detail-modal")
        self.assertContains(response, 'data-cashflow-kpi-detail="entradas"')
        self.assertContains(response, 'data-cashflow-kpi-detail="saidas"')
        self.assertContains(response, "Fluxo de caixa horizontal")
        self.assertContains(response, "fluxo-de-caixa/horizontal/")
        detalhes_previstos = response.context["fluxo_caixa"]["detalhes_previstos"]
        self.assertEqual(detalhes_previstos["entradas"][0]["data"], f"10/01/{ano_atual}")
        self.assertEqual(detalhes_previstos["entradas"][0]["nome"], "Cliente Critico")
        self.assertEqual(detalhes_previstos["entradas"][0]["categoria"], "Assinaturas")
        self.assertEqual(detalhes_previstos["entradas"][0]["valor_fmt"], "R$ 7.000,00")
        self.assertEqual(detalhes_previstos["saidas"][0]["data"], f"15/01/{ano_atual}")
        self.assertEqual(detalhes_previstos["saidas"][0]["nome"], "Fornecedor Critico")
        self.assertEqual(detalhes_previstos["saidas"][0]["categoria"], "Operacional")
        self.assertEqual(detalhes_previstos["saidas"][0]["valor_fmt"], "R$ 2.500,00")
        detalhes = response.context["fluxo_caixa"]["detalhes_lancamentos"][f"{ano_atual}-01"]
        self.assertEqual(detalhes["entradas"][0]["data"], f"20/01/{ano_atual}")
        self.assertEqual(detalhes["entradas"][0]["nome"], "Cliente Critico")
        self.assertEqual(detalhes["entradas"][0]["categoria"], "Assinaturas")
        self.assertEqual(detalhes["entradas"][0]["valor_fmt"], "R$ 900,00")
        self.assertEqual(detalhes["saidas"][0]["nome"], "Fornecedor Critico")
        response_horizontal = self.client.get(
            reverse(
                "dashboards:fluxo_caixa_horizontal",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {
                "modo": "diario",
                "ano": ano_atual,
                "mes": 1,
            },
        )

        self.assertEqual(response_horizontal.status_code, 200)
        horizontal = response_horizontal.context["fluxo_horizontal"]["modos"]
        self.assertIn("diario", horizontal)
        self.assertNotIn("semanal", horizontal)
        self.assertNotIn("anual", horizontal)
        self.assertEqual(horizontal["diario"]["periodos"][0]["rotulo"], f"01/01")
        self.assertEqual(horizontal["diario"]["receitas"][9]["previsao"], "R$ 7.000,00")
        self.assertEqual(horizontal["diario"]["receitas"][19]["realizado"], "R$ 900,00")
        self.assertEqual(horizontal["diario"]["despesas"][14]["previsao"], "R$ 2.500,00")
        self.assertEqual(horizontal["diario"]["despesas"][20]["realizado"], "R$ 300,00")
        self.assertContains(response_horizontal, "data-cashflow-horizontal-page")
        self.assertContains(response_horizontal, 'name="ano"')
        self.assertContains(response_horizontal, 'name="mes"')
        self.assertContains(response_horizontal, "data-horizontal-expand-all")
        self.assertContains(response_horizontal, "data-horizontal-collapse-all")
        self.assertContains(response_horizontal, "Saldo inicial")
        self.assertContains(response_horizontal, 'data-horizontal-parent="diario-saldo-root"')
        self.assertNotContains(response_horizontal, "app-sidebar")
        self.assertNotContains(response_horizontal, "app-shell")
        self.assertContains(response_horizontal, "Receitas recorrentes")
        self.assertContains(response_horizontal, "Assinaturas")
        self.assertContains(response_horizontal, "Despesas operacionais")
        self.assertContains(response_horizontal, "Operacional")
        self.assertContains(response_horizontal, 'data-horizontal-parent="diario-receitas-root"')
        self.assertContains(response_horizontal, "R$ 7.000,00")
        self.assertContains(response_horizontal, "R$ 900,00")

        response_anual = self.client.get(
            reverse(
                "dashboards:fluxo_caixa_horizontal",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {
                "modo": "anual",
                "ano": ano_atual,
            },
        )

        self.assertEqual(response_anual.status_code, 200)
        self.assertContains(response_anual, "Visao anual")
        self.assertContains(response_anual, 'data-cashflow-horizontal-month-filter hidden')

    def test_aprovacao_de_pagamentos_exibe_painel_e_modal_do_dia(self):
        hoje = date.today()
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        fornecedor = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=15001,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="Amazon AWS",
        )
        cliente = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=15002,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente recorrente",
        )
        fornecedor_anterior = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=15003,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="AWS anterior",
        )
        fornecedor_pago = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=15004,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="DATATEM SOLUCOES",
        )
        fornecedor_orfao = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=15005,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="Lancamento removido Omie",
        )
        categoria = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.03.01",
            descricao="Servicos em nuvem",
        )
        SincronizacaoOmie.objects.create(
            empresa=self.empresa,
            recurso="completa",
            status=SincronizacaoOmie.Status.CONCLUIDA,
            iniciada_em=timezone.now() - timedelta(hours=1),
            finalizada_em=timezone.now() - timedelta(minutes=55),
        )
        ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=601,
            descricao="Conta principal",
            saldo_atual=297182,
        )
        conta_atual = ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61001,
            fornecedor=fornecedor,
            categoria_principal=categoria,
            data_previsao=hoje,
            data_vencimento=hoje,
            valor_documento=12500,
            valor_a_pagar=12500,
            status_titulo="A VENCER",
        )
        AprovacaoPagamento.objects.create(
            conta_pagar=conta_atual,
            status=AprovacaoPagamento.Status.APROVADO,
            data_previsao_original=hoje,
            data_previsao_aprovada=hoje,
            aprovado_por=self.usuario,
            resposta_omie={"codigo_status": "0"},
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61002,
            fornecedor=fornecedor_anterior,
            categoria_principal=categoria,
            data_previsao=hoje - timedelta(days=7),
            data_vencimento=hoje - timedelta(days=3),
            valor_documento=9800,
            valor_a_pagar=9800,
            status_titulo="A VENCER",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61003,
            fornecedor=fornecedor_pago,
            categoria_principal=categoria,
            data_previsao=hoje,
            data_vencimento=hoje - timedelta(days=4),
            valor_documento=1646,
            valor_a_pagar=0,
            status_titulo=" pago ",
        )
        conta_orfa = ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61004,
            fornecedor=fornecedor_orfao,
            categoria_principal=categoria,
            data_previsao=hoje,
            data_vencimento=hoje,
            valor_documento=700,
            valor_a_pagar=700,
            status_titulo="A VENCER",
            ativo_omie=False,
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=62001,
            cliente=cliente,
            categoria_principal=categoria,
            data_vencimento=hoje,
            valor_documento=4300,
            valor_a_receber=4300,
            status_titulo="A VENCER",
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                    "dashboard_slug": "aprovacao-de-pagamentos",
                },
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Aprovacao de pagamentos")
        self.assertContains(response, "Abrir aprovacao de pagamentos")
        self.assertContains(response, "Consultar recebimentos do dia")
        self.assertContains(response, "Historico")
        self.assertContains(response, "Historico de aprovacao")
        self.assertContains(response, 'data-history-modal')
        self.assertContains(response, 'data-confirm-payment-history-export')
        self.assertContains(response, 'name="historico_inicio"')
        self.assertContains(response, 'name="historico_fim"')
        self.assertContains(response, "Amazon AWS")
        self.assertContains(response, "Servicos em nuvem")
        self.assertNotContains(response, "Lancamento removido Omie")
        pagamentos_contexto = response.context["aprovacao_pagamentos"]["pagamentos"]
        nomes_pagamentos = [pagamento["nome"] for pagamento in pagamentos_contexto]
        self.assertIn("Amazon AWS", nomes_pagamentos)
        self.assertNotIn("DATATEM SOLUCOES", nomes_pagamentos)
        self.assertContains(response, 'data-payment-value="12500.00"')
        self.assertContains(response, "Aprovar")
        self.assertContains(response, "Reagendar")
        self.assertContains(response, "Aprovar selecionados")
        self.assertContains(response, "Reagendar selecionados")
        self.assertContains(response, "Nova previsao")
        self.assertContains(response, "Previsao")
        self.assertContains(response, "Periodo do painel")
        self.assertNotContains(response, "Todos os projetos")
        self.assertNotContains(response, "Todos os departamentos")
        self.assertNotContains(response, "Limpar filtros")
        self.assertContains(response, "Categorias")
        self.assertContains(response, "Atualizar")
        self.assertContains(response, "Cliente recorrente")
        self.assertContains(response, 'data-payment-modal')
        self.assertContains(response, 'data-payment-success-modal')
        self.assertContains(response, 'data-payment-status-chart')
        self.assertContains(response, 'data-payment-status-detail-modal')
        self.assertContains(response, 'data-payment-status-detail-title')
        self.assertContains(response, "Operacao concluida")
        self.assertNotContains(response, "AWS anterior")

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                    "dashboard_slug": "aprovacao-de-pagamentos",
                },
            ),
            {
                "modal": "pagamentos",
                "aprovacao_inicio": (hoje - timedelta(days=7)).isoformat(),
                "aprovacao_fim": hoje.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "AWS anterior")
        self.assertContains(response, 'name="aprovacao_inicio"')
        self.assertContains(response, 'data-payment-open-initial="pagamentos"')
        pagamentos_contexto = response.context["aprovacao_pagamentos"]["pagamentos"]
        nomes_pagamentos = [pagamento["nome"] for pagamento in pagamentos_contexto]
        self.assertIn("AWS anterior", nomes_pagamentos)
        self.assertNotIn("DATATEM SOLUCOES", nomes_pagamentos)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                    "dashboard_slug": "aprovacao-de-pagamentos",
                },
            ),
            {
                "modal": "historico",
                "historico_inicio": (hoje - timedelta(days=7)).isoformat(),
                "historico_fim": hoje.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-payment-open-initial="historico"')
        self.assertContains(response, "AWS anterior")
        self.assertContains(response, "Aprovado")
        self.assertContains(response, "Pendente")

        response = self.client.get(
            reverse(
                "dashboards:exportar_historico_aprovacao_pagamentos",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            {
                "aprovacao_inicio": hoje.isoformat(),
                "aprovacao_fim": hoje.isoformat(),
                "historico_inicio": (hoje - timedelta(days=7)).isoformat(),
                "historico_fim": hoje.isoformat(),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(
            [cell.value for cell in worksheet[1]],
            ["Empresa", "Fornecedor / categoria", "Previsao", "Valor", "Status"],
        )
        self.assertEqual(worksheet["B2"].value, "AWS anterior\nServicos em nuvem")
        self.assertEqual(worksheet["E2"].value, "Pendente")

    @patch("apps.dashboards.aprovacao_pagamentos_services.consultar_conta_pagar")
    @patch("apps.dashboards.aprovacao_pagamentos_services.alterar_conta_pagar")
    def test_salvar_aprovacao_pagamentos_persiste_e_reagenda_na_omie(
        self,
        alterar_conta_pagar_mock,
        consultar_conta_pagar_mock,
    ):
        hoje = date.today()
        nova_previsao = hoje + timedelta(days=5)
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        integracao = IntegracaoOmie(empresa=self.empresa, app_key="app-key")
        integracao.definir_app_secret("app-secret")
        integracao.save()
        fornecedor = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=81001,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="Fornecedor lote",
        )
        conta_corrente = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=91001,
            descricao="Conta principal",
            saldo_atual=1000,
        )
        categoria = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.03.01",
            descricao="Servicos",
        )
        conta_aprovada = ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=71001,
            codigo_cliente_fornecedor=fornecedor.codigo_cliente_omie,
            fornecedor=fornecedor,
            id_conta_corrente=conta_corrente.codigo_omie,
            conta_corrente=conta_corrente,
            codigo_categoria=categoria.codigo,
            categoria_principal=categoria,
            data_previsao=hoje,
            data_vencimento=hoje,
            valor_documento=100,
            valor_a_pagar=100,
            status_titulo="A VENCER",
        )
        conta_reagendada = ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=71002,
            codigo_cliente_fornecedor=fornecedor.codigo_cliente_omie,
            fornecedor=fornecedor,
            id_conta_corrente=conta_corrente.codigo_omie,
            conta_corrente=conta_corrente,
            codigo_categoria=categoria.codigo,
            categoria_principal=categoria,
            data_previsao=hoje,
            data_vencimento=hoje + timedelta(days=2),
            valor_documento=250,
            valor_a_pagar=250,
            status_titulo="A VENCER",
        )
        alterar_conta_pagar_mock.return_value = {
            "codigo_status": "0",
            "descricao_status": "Lancamento alterado com sucesso!",
        }
        consultar_conta_pagar_mock.return_value = {
            "codigo_lancamento_omie": conta_aprovada.codigo_lancamento_omie,
            "codigo_cliente_fornecedor": fornecedor.codigo_cliente_omie,
            "data_vencimento": hoje.strftime("%d/%m/%Y"),
            "valor_documento": 100,
            "codigo_categoria": categoria.codigo,
            "data_previsao": hoje.strftime("%d/%m/%Y"),
            "id_conta_corrente": conta_corrente.codigo_omie,
            "observacao": "",
        }
        self.client.force_login(self.usuario)

        response = self.client.post(
            reverse(
                "dashboards:salvar_aprovacao_pagamentos",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            data=json.dumps(
                {
                    "itens": [
                        {"id": conta_aprovada.pk, "status": "approved"},
                        {
                            "id": conta_reagendada.pk,
                            "status": "rescheduled",
                            "new_date": nova_previsao.isoformat(),
                        },
                    ]
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        dados_resposta = response.json()
        self.assertIn(
            {
                "id": conta_aprovada.pk,
                "status": "approved",
                "data_previsao": hoje.isoformat(),
                "omie": "alterado",
            },
            dados_resposta["resultados"],
        )
        self.assertIn(
            {
                "id": conta_reagendada.pk,
                "status": "rescheduled",
                "data_previsao": nova_previsao.isoformat(),
                "omie": "alterado",
            },
            dados_resposta["resultados"],
        )
        conta_reagendada.refresh_from_db()
        aprovacao_aprovada = AprovacaoPagamento.objects.get(conta_pagar=conta_aprovada)
        aprovacao_reagendada = AprovacaoPagamento.objects.get(
            conta_pagar=conta_reagendada
        )
        self.assertEqual(conta_reagendada.data_previsao, nova_previsao)
        self.assertEqual(
            aprovacao_aprovada.status,
            AprovacaoPagamento.Status.APROVADO,
        )
        self.assertEqual(
            aprovacao_reagendada.status,
            AprovacaoPagamento.Status.REAGENDADO,
        )
        consultar_conta_pagar_mock.assert_called_once_with(
            integracao,
            conta_aprovada.codigo_lancamento_omie,
        )
        self.assertEqual(alterar_conta_pagar_mock.call_count, 2)
        _, payload_aprovacao = alterar_conta_pagar_mock.call_args_list[0].args
        self.assertEqual(payload_aprovacao["codigo_lancamento_omie"], 71001)
        self.assertEqual(payload_aprovacao["observacao"], "APROVADO")
        _, payload_reagendamento = alterar_conta_pagar_mock.call_args_list[1].args
        self.assertEqual(payload_reagendamento["codigo_lancamento_omie"], 71002)
        self.assertEqual(
            payload_reagendamento["data_previsao"],
            nova_previsao.strftime("%d/%m/%Y"),
        )
        conta_aprovada.refresh_from_db()
        self.assertEqual(conta_aprovada.dados_originais["observacao"], "APROVADO")

        alterar_conta_pagar_mock.reset_mock()
        consultar_conta_pagar_mock.reset_mock()
        response = self.client.post(
            reverse(
                "dashboards:salvar_aprovacao_pagamentos",
                kwargs={"empresa_slug": self.empresa.slug},
            ),
            data=json.dumps(
                {
                    "itens": [
                        {
                            "id": conta_reagendada.pk,
                            "status": "rescheduled",
                            "new_date": nova_previsao.isoformat(),
                        },
                    ]
                }
            ),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 207)
        self.assertEqual(
            response.json()["erros"][0]["erro"],
            "Lancamento ja enviado ao Omie. Altere somente no Omie.",
        )
        alterar_conta_pagar_mock.assert_not_called()
        consultar_conta_pagar_mock.assert_not_called()

    def test_fluxo_de_caixa_usa_apenas_lancamentos_realizados(self):
        ano_atual = date.today().year
        conta_visivel = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=301,
            descricao="Conta visivel",
            saldo_inicial=1000,
            saldo_atual=10,
        )
        conta_omitida = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=302,
            descricao="Conta omitida",
            saldo_inicial=9000,
            nao_fluxo=True,
            nao_resumo=True,
        )
        categoria_transferencia = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="9.99.01",
            descricao="Entrada de Transferencia",
            transferencia=True,
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=31001,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_previsao=date(ano_atual, 1, 10),
            data_vencimento=date(ano_atual, 1, 10),
            valor_documento=100,
            valor_a_receber=100,
            status_titulo="ATRASADO",
            dados_originais={"valor_iss": 10},
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=31003,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_previsao=date(ano_atual, 2, 10),
            data_vencimento=date(ano_atual, 2, 10),
            valor_documento=60,
            valor_a_receber=60,
            status_titulo="A VENCER",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=31004,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_previsao=date(ano_atual, 2, 11),
            data_vencimento=date(ano_atual, 2, 11),
            valor_documento=600,
            valor_a_receber=600,
            status_titulo="RECEBIDO",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=31005,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_previsao=date.today() + timedelta(days=10),
            data_vencimento=date.today() + timedelta(days=10),
            valor_documento=5000,
            valor_a_receber=5000,
            status_titulo="A VENCER",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=31002,
            id_conta_corrente=conta_omitida.codigo_omie,
            conta_corrente=conta_omitida,
            data_previsao=date(ano_atual, 1, 11),
            data_vencimento=date(ano_atual, 1, 11),
            valor_documento=900,
            valor_a_receber=900,
            status_titulo="ATRASADO",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=32001,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_entrada=date(ano_atual, 1, 1),
            data_previsao=date(ano_atual, 1, 15),
            data_vencimento=date(ano_atual, 1, 15),
            valor_documento=40,
            valor_a_pagar=40,
            status_titulo="ATRASADO",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=32002,
            id_conta_corrente=conta_omitida.codigo_omie,
            conta_corrente=conta_omitida,
            data_entrada=date(ano_atual, 1, 1),
            data_previsao=date(ano_atual, 1, 16),
            data_vencimento=date(ano_atual, 1, 16),
            valor_documento=400,
            valor_a_pagar=400,
            status_titulo="ATRASADO",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=32003,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_entrada=date(ano_atual, 1, 1),
            data_previsao=date(ano_atual, 2, 15),
            data_vencimento=date(ano_atual, 1, 16),
            valor_documento=70,
            valor_a_pagar=70,
            status_titulo="A VENCER",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=32004,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_entrada=date.today(),
            data_previsao=date.today() + timedelta(days=5),
            data_vencimento=date.today(),
            valor_documento=30,
            valor_a_pagar=30,
            status_titulo="A VENCER",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=32005,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_entrada=date.today(),
            data_previsao=date.today() + timedelta(days=10),
            data_vencimento=date.today() + timedelta(days=10),
            valor_documento=800,
            valor_a_pagar=800,
            status_titulo="A VENCER",
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=33001,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_lancamento=date(ano_atual, 1, 20),
            natureza="R",
            valor_lancamento=25,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=33004,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_lancamento=date(ano_atual, 1, 21),
            natureza="P",
            valor_lancamento=15,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=33002,
            codigo_conta_corrente=conta_omitida.codigo_omie,
            conta_corrente=conta_omitida,
            data_lancamento=date(ano_atual, 1, 21),
            natureza="P",
            valor_lancamento=250,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=33003,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            codigo_categoria=categoria_transferencia.codigo,
            categoria_principal=categoria_transferencia,
            data_lancamento=date(ano_atual, 1, 22),
            natureza="R",
            valor_lancamento=700,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=34001,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_previsao=date(ano_atual, 1, 10),
            data_vencimento=date(ano_atual, 1, 10),
            grupo="CONTA_A_RECEBER",
            natureza="R",
            status="ABERTO",
            valor_titulo=100,
            valor_liquido=100,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=34002,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_previsao=date(ano_atual, 2, 10),
            data_vencimento=date(ano_atual, 2, 10),
            grupo="CONTA_A_RECEBER",
            natureza="R",
            status="ABERTO",
            valor_titulo=60,
            valor_liquido=60,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=34003,
            codigo_conta_corrente=conta_omitida.codigo_omie,
            conta_corrente=conta_omitida,
            data_previsao=date(ano_atual, 1, 11),
            data_vencimento=date(ano_atual, 1, 11),
            grupo="CONTA_A_RECEBER",
            natureza="R",
            status="ABERTO",
            valor_titulo=900,
            valor_liquido=900,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=34004,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_previsao=date(ano_atual, 1, 20),
            data_pagamento=date(ano_atual, 1, 20),
            data_vencimento=date(ano_atual, 1, 20),
            grupo="CONTA_A_RECEBER",
            natureza="R",
            status="PAGO",
            valor_titulo=25,
            valor_liquido=25,
            valor_pago=25,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=34005,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            codigo_categoria=categoria_transferencia.codigo,
            categoria_principal=categoria_transferencia,
            data_previsao=date(ano_atual, 1, 22),
            data_pagamento=date(ano_atual, 1, 22),
            data_vencimento=date(ano_atual, 1, 22),
            grupo="CONTA_A_RECEBER",
            natureza="R",
            status="PAGO",
            valor_titulo=700,
            valor_liquido=700,
            valor_pago=700,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=34006,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_previsao=date(ano_atual, 1, 15),
            data_vencimento=date(ano_atual, 1, 15),
            grupo="CONTA_A_PAGAR",
            natureza="P",
            status="ABERTO",
            valor_titulo=40,
            valor_liquido=40,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=34007,
            codigo_conta_corrente=conta_omitida.codigo_omie,
            conta_corrente=conta_omitida,
            data_previsao=date(ano_atual, 1, 16),
            data_vencimento=date(ano_atual, 1, 16),
            grupo="CONTA_A_PAGAR",
            natureza="P",
            status="ABERTO",
            valor_titulo=400,
            valor_liquido=400,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=34008,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_previsao=date(ano_atual, 1, 21),
            data_vencimento=date(ano_atual, 1, 21),
            grupo="CONTA_A_PAGAR",
            natureza="P",
            status="PAGO",
            valor_titulo=15,
            valor_liquido=15,
            valor_pago=15,
        )

        contexto = fluxo_de_caixa(
            self.empresa,
            f"mes-{ano_atual}-01",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(contexto["entradas"], [25.0])
        self.assertEqual(contexto["saidas"], [15.0])
        self.assertEqual(contexto["saldo_acumulado"], [0.0])
        self.assertEqual(
            contexto["indicadores"][1]["valor_completo"],
            "R$ 90,00",
        )
        self.assertEqual(
            contexto["indicadores"][2]["valor_completo"],
            "R$ 40,00",
        )

    def test_fluxo_de_caixa_previstos_respeitam_fim_do_periodo_filtrado(self):
        conta = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=601,
            descricao="Conta principal",
            saldo_atual=1000,
        )
        cliente = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=6001,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Cliente ABC",
        )
        fornecedor = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=6002,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="Fornecedor ABC",
        )

        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61001,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            cliente=cliente,
            data_previsao=date(2026, 7, 20),
            data_vencimento=date(2026, 7, 20),
            valor_documento=100,
            valor_a_receber=100,
            status_titulo="ATRASADO",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61002,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            cliente=cliente,
            data_previsao=date(2026, 8, 31),
            data_vencimento=date(2026, 8, 31),
            valor_documento=200,
            valor_a_receber=200,
            status_titulo="A VENCER",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61003,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            cliente=cliente,
            data_previsao=date(2026, 9, 1),
            data_vencimento=date(2026, 9, 1),
            valor_documento=300,
            valor_a_receber=300,
            status_titulo="A VENCER",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61004,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            cliente=cliente,
            data_previsao=date(2026, 8, 15),
            data_vencimento=date(2026, 8, 15),
            valor_documento=400,
            valor_a_receber=400,
            status_titulo="RECEBIDO",
        )
        PesqTituloFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=61004,
            cliente_fornecedor=cliente,
            codigo_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            natureza="R",
            status="RECEBIDO",
            data_previsao=date(2026, 8, 15),
            data_vencimento=date(2026, 8, 15),
            valor_titulo=400,
            valor_aberto=40,
            valor_liquido=360,
            valor_pago=360,
        )

        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=62001,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            fornecedor=fornecedor,
            data_previsao=date(2026, 7, 25),
            data_vencimento=date(2026, 7, 25),
            valor_documento=50,
            valor_a_pagar=50,
            status_titulo="ATRASADO",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=62002,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            fornecedor=fornecedor,
            data_previsao=date(2026, 8, 31),
            data_vencimento=date(2026, 8, 31),
            valor_documento=70,
            valor_a_pagar=70,
            status_titulo="A VENCER",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=62003,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            fornecedor=fornecedor,
            data_previsao=date(2026, 9, 1),
            data_vencimento=date(2026, 9, 1),
            valor_documento=90,
            valor_a_pagar=90,
            status_titulo="A VENCER",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=62004,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            fornecedor=fornecedor,
            data_previsao=date(2026, 8, 15),
            data_vencimento=date(2026, 8, 15),
            valor_documento=300,
            valor_a_pagar=300,
            status_titulo="PAGO",
        )
        PesqTituloFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=62004,
            cliente_fornecedor=fornecedor,
            codigo_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            natureza="P",
            status="PAGO",
            data_previsao=date(2026, 8, 15),
            data_vencimento=date(2026, 8, 15),
            valor_titulo=300,
            valor_aberto=30,
            valor_liquido=270,
            valor_pago=270,
        )

        contexto = fluxo_de_caixa(
            self.empresa,
            "mes-2026-08",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(
            contexto["indicadores"][1]["valor_completo"],
            "R$ 340,00",
        )
        self.assertEqual(
            contexto["indicadores"][2]["valor_completo"],
            "R$ 150,00",
        )
        self.assertEqual(len(contexto["detalhes_previstos"]["entradas"]), 3)
        self.assertEqual(len(contexto["detalhes_previstos"]["saidas"]), 3)

    def test_fluxo_de_caixa_inclui_saldo_aberto_de_movimentos_parciais(self):
        conta = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=701,
            descricao="Conta principal",
            saldo_atual=1000,
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=71001,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_documento=10000,
            valor_a_pagar=10000,
            status_titulo="PAGO",
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=71001,
            codigo_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            grupo="CONTA_A_PAGAR",
            natureza="P",
            status="PAGO",
            liquidado=True,
            data_previsao=date.today(),
            data_pagamento=date.today(),
            data_vencimento=date.today(),
            valor_titulo=10000,
            valor_aberto=5000,
            valor_liquido=5000,
            valor_pago=5000,
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=71002,
            codigo_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            grupo="CONTA_A_PAGAR",
            natureza="P",
            status="PAGO",
            liquidado=True,
            data_previsao=date.today() + timedelta(days=40),
            data_pagamento=date.today() + timedelta(days=40),
            data_vencimento=date.today() + timedelta(days=40),
            valor_titulo=900,
            valor_aberto=900,
            valor_liquido=900,
            valor_pago=900,
        )

        contexto = fluxo_de_caixa(
            self.empresa,
            f"mes-{date.today().year}-{date.today().month:02d}",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(
            contexto["indicadores"][2]["valor_completo"],
            "R$ 5.000,00",
        )

    def test_fluxo_de_caixa_inclui_saldo_aberto_de_titulo_pago_pesquisado(self):
        conta = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=711,
            descricao="Conta principal",
            saldo_atual=1000,
        )
        fornecedor = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=9990499663,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="LEVISA",
        )
        categoria = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.05.02",
            descricao="Servicos Prestados",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=71003,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            fornecedor=fornecedor,
            categoria_principal=categoria,
            codigo_categoria=categoria.codigo,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_documento=6444,
            valor_a_pagar=6444,
            status_titulo="PAGO",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=71004,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            fornecedor=fornecedor,
            categoria_principal=categoria,
            codigo_categoria=categoria.codigo,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_documento=1200,
            valor_a_pagar=1200,
            status_titulo="ATRASADO",
        )
        PesqTituloFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=71003,
            codigo_cliente_fornecedor=fornecedor.codigo_cliente_omie,
            cliente_fornecedor=fornecedor,
            codigo_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            codigo_categoria=categoria.codigo,
            categoria_principal=categoria,
            natureza="P",
            status="PAGO",
            liquidado=False,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_titulo=6444,
            valor_aberto=1444,
            valor_liquido=5000,
            valor_pago=5000,
        )
        PesqTituloFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=71004,
            codigo_cliente_fornecedor=fornecedor.codigo_cliente_omie,
            cliente_fornecedor=fornecedor,
            codigo_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            codigo_categoria=categoria.codigo,
            categoria_principal=categoria,
            natureza="P",
            status="ATRASADO",
            liquidado=False,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_titulo=1200,
            valor_aberto=700,
            valor_liquido=500,
            valor_pago=500,
        )

        contexto = fluxo_de_caixa(
            self.empresa,
            f"mes-{date.today().year}-{date.today().month:02d}",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(
            contexto["indicadores"][2]["valor_completo"],
            "R$ 2.644,00",
        )
        self.assertEqual(len(contexto["detalhes_previstos"]["saidas"]), 2)
        self.assertIn(
            {
                "data": date.today().strftime("%d/%m/%Y"),
                "nome": "LEVISA",
                "categoria": "Servicos Prestados",
                "valor": 1444.0,
                "valor_fmt": "R$ 1.444,00",
            },
            contexto["detalhes_previstos"]["saidas"],
        )

    def test_fluxo_de_caixa_inclui_saldo_aberto_de_titulo_recebido_pesquisado(self):
        conta = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=712,
            descricao="Conta principal",
            saldo_atual=1000,
        )
        cliente = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=10453065883,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="ECOSYSTEM",
        )
        categoria = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.01.02",
            descricao="Servicos Prestados",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=72003,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            cliente=cliente,
            categoria_principal=categoria,
            codigo_categoria=categoria.codigo,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_documento=3000,
            valor_a_receber=3000,
            status_titulo="RECEBIDO",
        )
        PesqTituloFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=72003,
            codigo_cliente_fornecedor=cliente.codigo_cliente_omie,
            cliente_fornecedor=cliente,
            codigo_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            codigo_categoria=categoria.codigo,
            categoria_principal=categoria,
            natureza="R",
            status="RECEBIDO",
            liquidado=False,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_titulo=3000,
            valor_aberto=935.5,
            valor_liquido=2064.5,
            valor_pago=2064.5,
        )

        contexto = fluxo_de_caixa(
            self.empresa,
            f"mes-{date.today().year}-{date.today().month:02d}",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(
            contexto["indicadores"][1]["valor_completo"],
            "R$ 935,50",
        )
        self.assertIn(
            {
                "data": date.today().strftime("%d/%m/%Y"),
                "nome": "ECOSYSTEM",
                "categoria": "Servicos Prestados",
                "valor": 935.5,
                "valor_fmt": "R$ 935,50",
            },
            contexto["detalhes_previstos"]["entradas"],
        )

    def test_fluxo_de_caixa_deduplica_previstos_entre_movimentos_e_contas(self):
        conta = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=801,
            descricao="Conta principal",
            saldo_atual=1000,
        )
        fornecedor = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=8001,
            tipo=CadastroOmie.Tipo.FORNECEDOR,
            nome_fantasia="ALARMCEL",
        )
        categoria = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.05.02",
            descricao="Servicos Prestados",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=81001,
            id_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            fornecedor=fornecedor,
            categoria_principal=categoria,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_documento=566.68,
            valor_a_pagar=566.68,
            status_titulo="ATRASADO",
        )
        MovimentoFinanceiroOmie.objects.create(
            empresa=self.empresa,
            codigo_titulo=81001,
            codigo_conta_corrente=conta.codigo_omie,
            conta_corrente=conta,
            cliente_fornecedor=fornecedor,
            categoria_principal=categoria,
            grupo="CONTA_A_PAGAR",
            natureza="P",
            status="PAGO",
            liquidado=False,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_titulo=566.68,
            valor_aberto=566.68,
            valor_liquido=0,
            valor_pago=0,
        )

        contexto = fluxo_de_caixa(
            self.empresa,
            f"mes-{date.today().year}-{date.today().month:02d}",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(
            contexto["indicadores"][2]["valor_completo"],
            "R$ 566,68",
        )
        self.assertEqual(len(contexto["detalhes_previstos"]["saidas"]), 1)
        self.assertEqual(
            contexto["detalhes_previstos"]["saidas"][0]["nome"],
            "ALARMCEL",
        )

    def test_fluxo_de_caixa_usa_movimentos_em_vez_do_resumo_financeiro_omie(self):
        self.empresa.resumo_financeiro_omie = {
            "contaReceber": {
                "nTotal": 72,
                "vAtraso": 17971.01,
                "vTotal": 17971.01,
            },
            "contaPagar": {
                "nTotal": 61,
                "vAtraso": 741842.13,
                "vTotal": 759831.78,
            },
        }
        self.empresa.save(update_fields=["resumo_financeiro_omie"])
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=72001,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_documento=999999,
            valor_a_receber=999999,
            status_titulo="ABERTO",
        )
        ContaPagarOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=72002,
            data_previsao=date.today(),
            data_vencimento=date.today(),
            valor_documento=999999,
            valor_a_pagar=999999,
            status_titulo="ABERTO",
        )

        contexto = fluxo_de_caixa(
            self.empresa,
            f"mes-{date.today().year}-{date.today().month:02d}",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(
            contexto["indicadores"][1]["valor_completo"],
            "R$ 0,00",
        )
        self.assertEqual(
            contexto["indicadores"][2]["valor_completo"],
            "R$ 0,00",
        )

    def test_fluxo_de_caixa_reconstroi_saldo_inicial_por_movimentos_realizados(self):
        ano_atual = date.today().year
        ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=350,
            descricao="Conta corrente",
            saldo_atual=1000,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=35001,
            data_lancamento=date(ano_atual, 1, 10),
            natureza="R",
            valor_lancamento=300,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=35002,
            data_lancamento=date(ano_atual, 2, 10),
            natureza="P",
            valor_lancamento=120,
        )

        contexto = fluxo_de_caixa(
            self.empresa,
            f"mes-{ano_atual}-02",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(contexto["saldo_acumulado"], [1120.0])

    def test_fluxo_de_caixa_usa_saldo_atual_provisorio_das_contas_correntes(self):
        ano_atual = date.today().year
        self.empresa.saldo_contas_omie = Decimal("219720.53")
        self.empresa.save(update_fields=["saldo_contas_omie"])
        conta_visivel = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=601,
            descricao="Conta visivel",
            saldo_inicial=900,
            saldo_atual=Decimal("1234.56"),
            dados_originais={
                "extrato": {
                    "dPeriodoInicial": f"01/01/{ano_atual}",
                    "nSaldoAnterior": 16501.12,
                }
            },
        )
        conta_omitida = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=602,
            descricao="Conta omitida",
            saldo_inicial=5000,
            saldo_atual=Decimal("5000"),
            nao_fluxo=True,
            nao_resumo=True,
        )
        ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=603,
            descricao="Conta sem saldo provisorio",
            saldo_inicial=8000,
            dados_originais={
                "extrato": {
                    "dPeriodoInicial": f"01/01/{ano_atual}",
                    "nSaldoAnterior": 0,
                }
            },
        )
        categoria_transferencia = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="9.99.04",
            descricao="Entrada de Transferencia",
            transferencia=True,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61001,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_lancamento=date(ano_atual, 1, 10),
            natureza="R",
            valor_lancamento=100,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61002,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_lancamento=date(ano_atual, 1, 11),
            natureza="P",
            valor_lancamento=50,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61003,
            codigo_conta_corrente=conta_omitida.codigo_omie,
            conta_corrente=conta_omitida,
            data_lancamento=date(ano_atual, 1, 12),
            natureza="R",
            valor_lancamento=900,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=61004,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            codigo_categoria=categoria_transferencia.codigo,
            categoria_principal=categoria_transferencia,
            data_lancamento=date(ano_atual, 1, 13),
            natureza="R",
            valor_lancamento=800,
        )

        contexto = fluxo_de_caixa(
            self.empresa,
            f"mes-{ano_atual}-01",
            empresas_ids=[self.empresa.pk],
        )

        self.assertEqual(contexto["indicadores"][0]["valor_completo"], "R$ 1.234,56")
        self.assertEqual(contexto["saldo_acumulado"][0], 16501.12)

    def test_visao_geral_ignora_lancamentos_de_contas_fora_do_fluxo_e_resumo(self):
        ano_atual = date.today().year
        conta_visivel = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=401,
            descricao="Conta visivel",
        )
        conta_omitida = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=402,
            descricao="Conta omitida",
            nao_fluxo=True,
            nao_resumo=True,
        )
        categoria_transferencia = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="9.99.02",
            descricao="Saida de Transferencia",
            transferencia=True,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=41001,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            data_lancamento=date(ano_atual, 1, 10),
            natureza="R",
            valor_lancamento=100,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=41002,
            codigo_conta_corrente=conta_omitida.codigo_omie,
            conta_corrente=conta_omitida,
            data_lancamento=date(ano_atual, 1, 10),
            natureza="R",
            valor_lancamento=900,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=41003,
            codigo_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            codigo_categoria=categoria_transferencia.codigo,
            categoria_principal=categoria_transferencia,
            data_lancamento=date(ano_atual, 1, 10),
            natureza="R",
            valor_lancamento=800,
        )

        contexto = visao_geral_financeira(
            self.empresa,
            f"mes-{ano_atual}-01",
            empresas_ids=[self.empresa.pk],
            regime_financeiro="caixa",
        )

        self.assertEqual(contexto["recebimentos"], [100.0])

    def test_dre_gerencial_ignora_titulos_de_contas_fora_do_fluxo_e_resumo(self):
        ano_atual = date.today().year
        conta_dre = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receita Bruta",
            ordem=1,
        )
        vendas = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Vendas",
            conta_pai=conta_dre,
            ordem=1,
        )
        categoria = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.05.01",
            descricao="Vendas",
            conta_dre=vendas,
        )
        categoria_transferencia = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="9.99.03",
            descricao="Entrada de Transferencia",
            transferencia=True,
            conta_dre=vendas,
        )
        conta_visivel = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=501,
            descricao="Conta visivel",
        )
        conta_omitida = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=502,
            descricao="Conta omitida",
            nao_fluxo=True,
            nao_resumo=True,
        )
        conta_adiantamento = ContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_omie=503,
            descricao=" Adiantamento de Cliente ",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=51001,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            categoria_principal=categoria,
            data_registro=date(ano_atual, 1, 10),
            valor_documento=100,
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=51002,
            id_conta_corrente=conta_omitida.codigo_omie,
            conta_corrente=conta_omitida,
            categoria_principal=categoria,
            data_registro=date(ano_atual, 1, 11),
            valor_documento=900,
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=51003,
            id_conta_corrente=conta_visivel.codigo_omie,
            conta_corrente=conta_visivel,
            codigo_categoria=categoria_transferencia.codigo,
            categoria_principal=categoria_transferencia,
            data_registro=date(ano_atual, 1, 12),
            valor_documento=800,
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=51004,
            id_conta_corrente=conta_adiantamento.codigo_omie,
            conta_corrente=conta_adiantamento,
            categoria_principal=categoria,
            data_registro=date(ano_atual, 1, 13),
            valor_documento=700,
        )

        contexto = dre_gerencial(
            self.empresa,
            f"mes-{ano_atual}-01",
            empresas_ids=[self.empresa.pk],
            regime_financeiro="competencia",
        )
        linha_vendas = next(linha for linha in contexto["linhas"] if linha["nome"] == "Vendas")

        self.assertEqual(linha_vendas["total"], Decimal("100"))

    def test_dre_gerencial_calcula_contas_de_resultado_em_cascata(self):
        ano_atual = date.today().year
        receita = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Receita Bruta",
            ordem=1,
        )
        vendas = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Vendas",
            conta_pai=receita,
            ordem=1,
        )
        deducoes = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Deducoes",
            sinal=ContaDRE.Sinal.SUBTRACAO,
            ordem=2,
        )
        impostos = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Impostos",
            conta_pai=deducoes,
            sinal=ContaDRE.Sinal.SUBTRACAO,
            ordem=1,
        )
        resultado_bruto = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Resultado bruto",
            sinal=ContaDRE.Sinal.RESULTADO,
            ordem=3,
        )
        despesas = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Despesas",
            sinal=ContaDRE.Sinal.SUBTRACAO,
            ordem=4,
        )
        administrativas = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Administrativas",
            conta_pai=despesas,
            sinal=ContaDRE.Sinal.SUBTRACAO,
            ordem=1,
        )
        resultado_final = ContaDRE.objects.create(
            empresa=self.empresa,
            nome="Resultado final",
            sinal=ContaDRE.Sinal.RESULTADO,
            ordem=5,
        )
        categoria_vendas = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.10.01",
            descricao="Vendas",
            conta_dre=vendas,
        )
        categoria_impostos = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.10.01",
            descricao="Impostos",
            conta_dre=impostos,
        )
        categoria_despesas = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="2.20.01",
            descricao="Administrativas",
            conta_dre=administrativas,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=52001,
            categoria_principal=categoria_vendas,
            data_lancamento=date(ano_atual, 1, 10),
            valor_lancamento=1000,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=52002,
            categoria_principal=categoria_impostos,
            data_lancamento=date(ano_atual, 1, 11),
            valor_lancamento=200,
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=52003,
            categoria_principal=categoria_despesas,
            data_lancamento=date(ano_atual, 1, 12),
            valor_lancamento=300,
        )

        contexto = dre_gerencial(
            self.empresa,
            f"mes-{ano_atual}-01",
            empresas_ids=[self.empresa.pk],
            regime_financeiro="caixa",
        )

        linhas = {linha["nome"]: linha for linha in contexto["linhas"]}
        self.assertEqual(linhas[resultado_bruto.nome]["total"], Decimal("800"))
        self.assertEqual(linhas[resultado_final.nome]["total"], Decimal("500"))

    def test_inadimplencia_exibe_indicadores_graficos_e_top_devedores(self):
        ano_atual = date.today().year
        EmpresaUsuario.objects.create(empresa=self.empresa, usuario=self.usuario)
        cliente_a = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=14001,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Devedor Alfa",
        )
        cliente_b = CadastroOmie.objects.create(
            empresa=self.empresa,
            codigo_cliente_omie=14002,
            tipo=CadastroOmie.Tipo.CLIENTE,
            nome_fantasia="Devedor Beta",
        )
        categoria = CategoriaOmie.objects.create(
            empresa=self.empresa,
            codigo="1.04.01",
            descricao="Servicos em atraso",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=15001,
            cliente=cliente_a,
            categoria_principal=categoria,
            data_emissao=date(ano_atual, 1, 1),
            data_vencimento=date(ano_atual, 1, 10),
            valor_documento=8000,
            valor_a_receber=8000,
            status_titulo="ATRASADO",
        )
        ContaReceberOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=15002,
            cliente=cliente_b,
            categoria_principal=categoria,
            data_emissao=date(ano_atual, 1, 5),
            data_vencimento=date(ano_atual, 2, 15),
            valor_documento=3000,
            valor_a_receber=3000,
            status_titulo="ATRASADO",
        )
        LancamentoContaCorrenteOmie.objects.create(
            empresa=self.empresa,
            codigo_lancamento_omie=16001,
            cliente_fornecedor=cliente_a,
            data_lancamento=date(ano_atual, 1, 20),
            natureza="R",
            valor_lancamento=1200,
        )
        self.client.force_login(self.usuario)

        response = self.client.get(
            reverse(
                "dashboards:dashboard",
                kwargs={
                    "empresa_slug": self.empresa.slug,
                    "area_slug": "financeiro",
                    "dashboard_slug": "inadimplencia",
                },
            ),
            {
                "_filtrar": "1",
                "periodo": f"ano-{ano_atual}",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Total inadimplente")
        self.assertContains(response, "% Inadimplencia")
        self.assertContains(response, "DSO")
        self.assertContains(response, "Recuperado no mes")
        self.assertContains(response, "Aging Schedule")
        self.assertContains(response, "% Inadimplencia")
        self.assertContains(response, "Devedor Alfa")
        self.assertContains(response, "Devedor Beta")
        self.assertContains(response, "data-aging-chart")
        self.assertContains(response, "data-delinquency-trend-chart")
